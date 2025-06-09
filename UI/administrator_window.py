import os
import sys
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill


import requests
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFrame, QStackedWidget, QSizePolicy,
    QDialog, QTableWidget, QTableWidgetItem, QHeaderView,
    QListWidget, QListWidgetItem, QTextEdit, QLineEdit,
    QComboBox, QDateEdit, QFormLayout, QMessageBox, QScrollArea, QFileDialog, QGroupBox, QSplitter
)
from PySide6.QtCore import Qt, QSize, QTimer, QDate
from PySide6.QtGui import QFont, QIcon, QPixmap, QImage, QColor


class AdminApp(QMainWindow):
    def __init__(self, user_id):
        super().__init__()
        self.user_id = user_id
        self.setWindowTitle("Гусарочка - Администратор")
        self.setMinimumSize(1280, 900)
        self.base_url = "http://localhost:8000"  # Замените на ваш базовый URL API

        # Инициализируем словарь для хранения карточек статистики
        self.stat_cards = {
            "users": None,
            "products": None,
            "orders": None,
            "returns": None
        }

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Sidebar
        self.create_admin_sidebar(main_layout)

        # Content area
        self.content_stack = QStackedWidget()

        # Dashboard page
        self.dashboard_page = self.create_dashboard_page()
        self.content_stack.addWidget(self.dashboard_page)

        # Users management page
        self.users_page = self.create_users_page()
        self.content_stack.addWidget(self.users_page)

        # Products management page
        self.products_page = self.create_products_page()
        self.content_stack.addWidget(self.products_page)

        # Orders page
        self.orders_page = self.create_orders_page()
        self.content_stack.addWidget(self.orders_page)

        # Returns page
        self.returns_page = self.create_returns_page()
        self.content_stack.addWidget(self.returns_page)

        # Reports page
        self.reports_page = self.create_reports_page()
        self.content_stack.addWidget(self.reports_page)

        main_layout.addWidget(self.content_stack, stretch=5)

        # Apply styles
        self.apply_admin_styles()

    def create_admin_sidebar(self, layout):
        sidebar = QFrame()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(280)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(20, 40, 20, 40)
        sidebar_layout.setSpacing(20)

        # Logo
        logo = QLabel("ГУСАРОЧКА\nАДМИНИСТРАТОР")
        logo.setFont(QFont('Montserrat', 18, QFont.Weight.Bold))
        logo.setStyleSheet("color: #d4af37;")
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sidebar_layout.addWidget(logo)

        # Navigation buttons
        self.dashboard_btn = QPushButton("Дашборд")
        self.dashboard_btn.setIcon(QIcon("icons/home.png"))
        self.dashboard_btn.setIconSize(QSize(24, 24))
        self.dashboard_btn.setObjectName("navButton")
        self.dashboard_btn.clicked.connect(lambda: self.content_stack.setCurrentIndex(0))

        self.users_btn = QPushButton("Пользователи")
        self.users_btn.setIcon(QIcon("icons/users.png"))
        self.users_btn.setIconSize(QSize(24, 24))
        self.users_btn.setObjectName("navButton")
        self.users_btn.clicked.connect(lambda: self.content_stack.setCurrentIndex(1))

        self.products_btn = QPushButton("Товары")
        self.products_btn.setIcon(QIcon("icons/products.png"))
        self.products_btn.setIconSize(QSize(24, 24))
        self.products_btn.setObjectName("navButton")
        self.products_btn.clicked.connect(lambda: self.content_stack.setCurrentIndex(2))

        self.orders_btn = QPushButton("Заказы")
        self.orders_btn.setIcon(QIcon("icons/orders.png"))
        self.orders_btn.setIconSize(QSize(24, 24))
        self.orders_btn.setObjectName("navButton")
        self.orders_btn.clicked.connect(lambda: self.content_stack.setCurrentIndex(3))

        self.returns_btn = QPushButton("Возвраты")
        self.returns_btn.setIcon(QIcon("icons/returns.png"))
        self.returns_btn.setIconSize(QSize(24, 24))
        self.returns_btn.setObjectName("navButton")
        self.returns_btn.clicked.connect(lambda: self.content_stack.setCurrentIndex(4))

        self.reports_btn = QPushButton("Отчеты")
        self.reports_btn.setIcon(QIcon("icons/reports.png"))
        self.reports_btn.setIconSize(QSize(24, 24))
        self.reports_btn.setObjectName("navButton")
        self.reports_btn.clicked.connect(lambda: self.content_stack.setCurrentIndex(5))

        sidebar_layout.addWidget(self.dashboard_btn)
        sidebar_layout.addWidget(self.users_btn)
        sidebar_layout.addWidget(self.products_btn)
        sidebar_layout.addWidget(self.orders_btn)
        sidebar_layout.addWidget(self.returns_btn)
        sidebar_layout.addWidget(self.reports_btn)
        sidebar_layout.addStretch()

        # Logout button
        logout_btn = QPushButton("Выйти")
        logout_btn.setObjectName("logoutButton")
        logout_btn.clicked.connect(self.close)
        sidebar_layout.addWidget(logout_btn)

        layout.addWidget(sidebar)

    def log_action(self, action_text):
        """Записывает действие в список последних действий"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {action_text}"

        # Находим виджет списка действий на дашборде
        dashboard_page = self.content_stack.widget(0)
        activity_list = dashboard_page.findChild(QListWidget, "activityList")

        if activity_list:
            # Если список пустой или содержит заглушки, очищаем его
            if activity_list.count() == 1 and "Нет последних действий" in activity_list.item(0).text():
                activity_list.clear()

            # Добавляем новое действие в начало списка
            activity_list.insertItem(0, log_entry)

            # Ограничиваем количество записей (например, последние 50 действий)
            if activity_list.count() > 50:
                activity_list.takeItem(activity_list.count() - 1)

    def create_dashboard_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(30)

        # Welcome section
        welcome = QLabel("Административная панель")
        welcome.setFont(QFont('Montserrat', 24, QFont.Weight.Bold))
        welcome.setStyleSheet("color: gold;")
        layout.addWidget(welcome)

        # Stats cards
        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(20)

        # Create stat cards (без иконок)
        stats = [
            {"id": "users", "title": "Пользователи", "value": "0"},
            {"id": "products", "title": "Товары", "value": "0"},
            {"id": "orders", "title": "Заказы", "value": "0"},
            {"id": "returns", "title": "Возвраты", "value": "0"}
        ]

        for stat in stats:
            card = QFrame()
            card.setObjectName("statCard")
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(20, 20, 20, 20)
            card_layout.setSpacing(10)

            # Только заголовок (без иконки)
            title = QLabel(stat["title"])
            title.setFont(QFont('Montserrat', 16))
            title.setStyleSheet("color: white;")
            title.setAlignment(Qt.AlignmentFlag.AlignCenter)  # Центрируем заголовок

            # Значение
            value = QLabel(stat["value"])
            value.setFont(QFont('Montserrat', 42, QFont.Weight.Bold))
            value.setStyleSheet("color: #d4af37;")
            value.setAlignment(Qt.AlignmentFlag.AlignCenter)

            card_layout.addWidget(title)
            card_layout.addWidget(value)
            card_layout.addStretch()

            stats_layout.addWidget(card)
            self.stat_cards[stat["id"]] = value

        layout.addLayout(stats_layout)

        # Остальной код остается без изменений...
        activity_label = QLabel("Последние действия")
        activity_label.setFont(QFont('Montserrat', 18, QFont.Weight.Bold))
        activity_label.setStyleSheet("color: white;")
        layout.addWidget(activity_label)

        activity_list = QListWidget()
        activity_list.setObjectName("activityList")
        activity_list.addItem("Нет последних действий")  # Начальное состояние
        layout.addWidget(activity_list)

        layout.addStretch()

        return page

    def update_dashboard_stats(self, users_count=None, products_count=None, orders_count=None, returns_count=None):
        """Обновляет статистику на дашборде"""
        if users_count is not None and self.stat_cards["users"]:
            self.stat_cards["users"].setText(str(users_count))

        if products_count is not None and self.stat_cards["products"]:
            self.stat_cards["products"].setText(str(products_count))

        if orders_count is not None and self.stat_cards["orders"]:
            self.stat_cards["orders"].setText(str(orders_count))

        if returns_count is not None and self.stat_cards["returns"]:
            self.stat_cards["returns"].setText(str(returns_count))

    def create_users_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        header = QHBoxLayout()

        title = QLabel("Управление пользователями")
        title.setFont(QFont('Montserrat', 20, QFont.Weight.Bold))
        title.setStyleSheet("color: gold;")

        header.addWidget(title)
        header.addStretch()
        layout.addLayout(header)

        # Search bar - новая улучшенная версия
        search_frame = QFrame()
        search_frame.setStyleSheet("background-color: #252525; border-radius: 5px;")
        search_layout = QHBoxLayout(search_frame)
        search_layout.setContentsMargins(15, 10, 15, 10)
        search_layout.setSpacing(10)

        self.user_search = QLineEdit()
        self.user_search.setPlaceholderText("Поиск по имени или роли...")
        self.user_search.setStyleSheet("""
            QLineEdit {
                background-color: #333;
                color: white;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
            }
        """)

        search_btn = QPushButton("Поиск")
        search_btn.setObjectName("searchButton")
        search_btn.setFixedWidth(100)
        search_btn.clicked.connect(self.search_users)

        search_layout.addWidget(self.user_search)
        search_layout.addWidget(search_btn)
        layout.addWidget(search_frame)

        # Users table
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(4)  # ID, Имя, Роль, Дата регистрации
        self.users_table.setHorizontalHeaderLabels(["ID", "Имя пользователя", "Роль", "Дата регистрации"])
        self.users_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.users_table.verticalHeader().setVisible(False)
        self.users_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.users_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.users_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)

        # Загружаем пользователей
        self.load_users()
        layout.addWidget(self.users_table, stretch=1)

        # Кнопки управления
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)

        add_user_btn = QPushButton("Добавить пользователя")
        add_user_btn.setObjectName("addButton")
        add_user_btn.clicked.connect(self.show_add_user_dialog)

        self.edit_user_btn = QPushButton("Редактировать")
        self.edit_user_btn.setObjectName("editButton")
        self.edit_user_btn.clicked.connect(self.edit_selected_user)
        self.edit_user_btn.setEnabled(False)

        self.delete_user_btn = QPushButton("Удалить")
        self.delete_user_btn.setObjectName("deleteButton")
        self.delete_user_btn.clicked.connect(self.delete_selected_user)
        self.delete_user_btn.setEnabled(False)

        refresh_btn = QPushButton("Обновить")
        refresh_btn.setObjectName("refreshButton")
        refresh_btn.clicked.connect(self.load_users)

        buttons_layout.addWidget(add_user_btn)
        buttons_layout.addWidget(self.edit_user_btn)
        buttons_layout.addWidget(self.delete_user_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(refresh_btn)
        layout.addLayout(buttons_layout)

        # Подключаем сигнал изменения выделения
        self.users_table.itemSelectionChanged.connect(self.update_buttons_state)

        return page

    def update_buttons_state(self):
        """Обновляет состояние кнопок в зависимости от выбранной строки"""
        selected = self.users_table.selectionModel().hasSelection()
        self.edit_user_btn.setEnabled(selected)
        self.delete_user_btn.setEnabled(selected)

    def edit_selected_user(self):
        """Редактирование выбранного пользователя"""
        selected_row = self.users_table.currentRow()
        if selected_row >= 0:
            user_id = int(self.users_table.item(selected_row, 0).text())
            username = self.users_table.item(selected_row, 1).text()
            role = self.users_table.item(selected_row, 2).text()
            created_at = self.users_table.item(selected_row, 3).text()

            user_data = {
                'user_id': user_id,
                'username': username,
                'role': role,
                'created_at': created_at
            }

            self.show_edit_user_dialog(user_data)

    def delete_selected_user(self):
        """Удаление выбранного пользователя"""
        selected_row = self.users_table.currentRow()
        if selected_row >= 0:
            user_id = int(self.users_table.item(selected_row, 0).text())
            self.delete_user(user_id)

    def search_users(self):
        search_text = self.user_search.text().strip().lower()
        self.log_action(f"Поиск пользователей: '{search_text}'")

        try:
            response = requests.get(f"{self.base_url}/users")
            if response.status_code == 200:
                users = response.json()
                filtered_users = []

                for user in users:
                    matches = [
                        search_text in user.get('username', '').lower(),
                        search_text in user.get('role', '').lower()
                    ]

                    if any(matches):
                        filtered_users.append(user)

                self.populate_users_table(filtered_users)
                self.log_action(f"Найдено {len(filtered_users)} пользователей")
            else:
                self.log_action(f"Ошибка поиска пользователей: код {response.status_code}")
                QMessageBox.warning(self, "Ошибка", "Не удалось выполнить поиск")
        except Exception as e:
            self.log_action(f"Ошибка при поиске пользователей: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при поиске: {str(e)}")

    def load_users(self):
        try:
            self.log_action("Запрос списка пользователей")
            response = requests.get(f"{self.base_url}/users")

            if response.status_code == 200:
                users = response.json()
                self.populate_users_table(users)
                self.update_dashboard_stats(len(users))
                self.log_action(f"Успешно загружено {len(users)} пользователей")
            else:
                self.log_action(f"Ошибка загрузки пользователей: код {response.status_code}")
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить пользователей. Код ошибки: {response.status_code}")
        except Exception as e:
            self.log_action(f"Ошибка при загрузке пользователей: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке пользователей: {str(e)}")

    def populate_users_table(self, users):
        self.users_table.setRowCount(0)

        for user in users:
            row_position = self.users_table.rowCount()
            self.users_table.insertRow(row_position)

            self.users_table.setItem(row_position, 0, QTableWidgetItem(str(user['user_id'])))
            self.users_table.setItem(row_position, 1, QTableWidgetItem(user['username']))

            # Отображаем роль как есть (ожидаем русские названия с сервера)
            role = user.get('role', 'Клиент')
            self.users_table.setItem(row_position, 2, QTableWidgetItem(role))

            # Форматирование даты
            try:
                dt = datetime.fromisoformat(user['created_at'])
                self.users_table.setItem(row_position, 3, QTableWidgetItem(dt.strftime("%d.%m.%Y %H:%M")))
            except (ValueError, KeyError):
                self.users_table.setItem(row_position, 3, QTableWidgetItem(user.get('created_at', '')))

    def delete_user(self, user_id):
        msg_box = QMessageBox()
        msg_box.setWindowTitle("Подтверждение удаления")
        msg_box.setText(f"Вы уверены, что хотите удалить пользователя с ID {user_id}?")
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #1a1a1a;
            }
            QLabel {
                color: white;
            }
            QPushButton {
                color: white;
                background-color: #333;
                padding: 5px 10px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #444;
            }
        """)
        msg_box.button(QMessageBox.StandardButton.Yes).setText("Да")
        msg_box.button(QMessageBox.StandardButton.No).setText("Нет")

        reply = msg_box.exec()

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.log_action(f"Удаление пользователя ID: {user_id}")
                response = requests.delete(f"{self.base_url}/users/{user_id}")

                if response.status_code == 200:
                    self.log_action(f"Пользователь ID: {user_id} успешно удален")
                    QMessageBox.information(self, "Успех", "Пользователь успешно удален")
                    self.load_users()
                else:
                    self.log_action(f"Ошибка удаления пользователя ID: {user_id}: {response.text}")
                    QMessageBox.warning(self, "Ошибка", f"Не удалось удалить пользователя: {response.text}")
            except Exception as e:
                self.log_action(f"Ошибка при удалении пользователя ID: {user_id}: {str(e)}")
                QMessageBox.critical(self, "Ошибка", f"Ошибка соединения: {str(e)}")

    def show_edit_user_dialog(self, user):
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактировать пользователя")
        dialog.setFixedSize(400, 300)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        title = QLabel(f"Редактирование пользователя {user['username']}")
        title.setFont(QFont('Montserrat', 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #d4af37;")
        layout.addWidget(title)

        # Form
        form_layout = QFormLayout()
        form_layout.setHorizontalSpacing(15)
        form_layout.setVerticalSpacing(10)

        self.edit_username_input = QLineEdit(user['username'])
        self.edit_username_input.setPlaceholderText("Имя пользователя")

        self.edit_password_input = QLineEdit()
        self.edit_password_input.setPlaceholderText("Новый пароль (оставьте пустым, чтобы не менять)")
        self.edit_password_input.setEchoMode(QLineEdit.Password)

        form_layout.addRow("Имя пользователя:", self.edit_username_input)
        form_layout.addRow("Новый пароль:", self.edit_password_input)

        layout.addLayout(form_layout)

        # Buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setObjectName("cancelButton")
        cancel_btn.clicked.connect(dialog.reject)

        save_btn = QPushButton("Сохранить")
        save_btn.setObjectName("saveButton")
        save_btn.clicked.connect(lambda: self.update_user(user['user_id'], dialog))

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(save_btn)
        layout.addLayout(buttons_layout)

        # Status
        self.edit_user_status = QLabel()
        self.edit_user_status.setWordWrap(True)
        layout.addWidget(self.edit_user_status)

        dialog.setStyleSheet("""
            QLineEdit, QComboBox {
                padding: 10px;
                font-size: 14px;
                border: 1px solid #444;
                border-radius: 4px;
                background: #333;
                color: white;
            }
            QLabel {
                color: white;
            }
            #saveButton {
                background-color: #d4af37;
                color: black;
                font-size: 16px;
                padding: 12px;
                border-radius: 5px;
            }
            #saveButton:hover {
                background-color: #f0d87c;
            }
            #cancelButton {
                background-color: #333;
                color: white;
                font-size: 16px;
                padding: 12px;
                border-radius: 5px;
            }
            #cancelButton:hover {
                background-color: #444;
            }
        """)

        dialog.exec()

    def update_user(self, user_id, dialog):
        username = self.edit_username_input.text().strip()
        password = self.edit_password_input.text().strip()
        self.log_action(f"Обновление пользователя ID: {user_id}")

        if not username:
            self.log_action("Ошибка: имя пользователя не может быть пустым")
            self.edit_user_status.setText("Имя пользователя не может быть пустым")
            self.edit_user_status.setStyleSheet("color: #ff5555;")
            return

        try:
            data = {
                "username": username
            }

            if password:
                data["password"] = password

            response = requests.put(f"{self.base_url}/users/{user_id}", json=data)

            if response.status_code == 200:
                self.log_action(f"Пользователь ID: {user_id} успешно обновлен")
                self.edit_user_status.setText("Пользователь успешно обновлен")
                self.edit_user_status.setStyleSheet("color: #55ff55;")
                QTimer.singleShot(1500, dialog.close)
                self.load_users()
            else:
                self.log_action(f"Ошибка обновления пользователя ID: {user_id}: {response.text}")
                self.edit_user_status.setText(f"Ошибка при обновлении пользователя: {response.text}")
                self.edit_user_status.setStyleSheet("color: #ff5555;")
        except Exception as e:
            self.log_action(f"Ошибка при обновлении пользователя ID: {user_id}: {str(e)}")
            self.edit_user_status.setText(f"Ошибка: {str(e)}")
            self.edit_user_status.setStyleSheet("color: #ff5555;")

    def create_products_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        header = QHBoxLayout()

        title = QLabel("Управление товарами")
        title.setFont(QFont('Montserrat', 20, QFont.Weight.Bold))
        title.setStyleSheet("color: gold;")

        header.addWidget(title)
        header.addStretch()
        layout.addLayout(header)

        # Search bar - теперь на всю ширину
        search_frame = QFrame()
        search_frame.setStyleSheet("background-color: #252525; border-radius: 5px;")
        search_layout = QHBoxLayout(search_frame)
        search_layout.setContentsMargins(15, 10, 15, 10)
        search_layout.setSpacing(10)

        self.product_search = QLineEdit()
        self.product_search.setPlaceholderText("Поиск по названию или артикулу...")  # Обновленный текст
        self.product_search.setStyleSheet("""
            QLineEdit {
                background-color: #333;
                color: white;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
            }
        """)

        search_btn = QPushButton("Поиск")
        search_btn.setObjectName("searchButton")
        search_btn.setFixedWidth(100)
        search_btn.clicked.connect(self.search_products)

        search_layout.addWidget(self.product_search)
        search_layout.addWidget(search_btn)
        layout.addWidget(search_frame)

        # Products table
        self.products_table = QTableWidget()
        self.products_table.setColumnCount(12)
        self.products_table.setHorizontalHeaderLabels([
            "ID", "Название", "Артикул", "Тип", "Материал",
            "Вставка", "Вес", "Цена", "На складе",
            "ID изображения", "Дата создания", "Изображение"
        ])
        self.products_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.products_table.verticalHeader().setVisible(False)
        self.products_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.products_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.products_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)

        # Загружаем товары
        self.load_products()
        layout.addWidget(self.products_table, stretch=1)  # Таблица теперь растягивается

        # Кнопки управления
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)

        self.add_product_btn = QPushButton("Добавить товар")
        self.add_product_btn.setObjectName("addButton")
        self.add_product_btn.clicked.connect(self.show_add_product_dialog)

        self.edit_product_btn = QPushButton("Редактировать")
        self.edit_product_btn.setObjectName("editButton")
        self.edit_product_btn.clicked.connect(self.edit_selected_product)
        self.edit_product_btn.setEnabled(False)

        self.delete_product_btn = QPushButton("Удалить")
        self.delete_product_btn.setObjectName("deleteButton")
        self.delete_product_btn.clicked.connect(self.delete_selected_product)
        self.delete_product_btn.setEnabled(False)

        refresh_btn = QPushButton("Обновить")
        refresh_btn.setObjectName("refreshButton")
        refresh_btn.clicked.connect(self.load_products)

        buttons_layout.addWidget(self.add_product_btn)
        buttons_layout.addWidget(self.edit_product_btn)
        buttons_layout.addWidget(self.delete_product_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(refresh_btn)
        layout.addLayout(buttons_layout)

        # Подключаем сигнал изменения выделения
        self.products_table.itemSelectionChanged.connect(self.update_product_buttons_state)

        return page

    def show_add_product_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить товар")
        dialog.setFixedSize(600, 800)  # Увеличенный размер для лучшего размещения

        # Основной стиль диалога
        dialog.setStyleSheet("""
            QDialog {
                background-color: #1a1a1a;
                color: white;
                font-family: 'Segoe UI';
            }
            QLabel {
                color: white;
            }
            QLineEdit, QTextEdit {
                background-color: #333;
                color: white;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
                selection-background-color: #d4af37;
                selection-color: black;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 14px;
                min-width: 100px;
            }
            #saveButton {
                background-color: #d4af37;
                color: black;
            }
            #saveButton:hover {
                background-color: #f0d87c;
            }
            #cancelButton {
                background-color: #444;
                color: white;
            }
            #cancelButton:hover {
                background-color: #555;
            }
            #browseButton {
                background-color: #555;
                color: white;
            }
            #browseButton:hover {
                background-color: #666;
            }
            QGroupBox {
                color: #d4af37;
                border: 1px solid #444;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)

        main_layout = QVBoxLayout(dialog)
        main_layout.setContentsMargins(25, 25, 25, 25)
        main_layout.setSpacing(15)

        # Заголовок
        title = QLabel("Добавление нового товара")
        title.setFont(QFont('Montserrat', 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #d4af37; margin-bottom: 15px;")
        main_layout.addWidget(title)

        # Основная форма с полями ввода
        form_scroll = QScrollArea()
        form_scroll.setWidgetResizable(True)
        form_scroll.setStyleSheet("border: none;")

        form_widget = QWidget()
        form_layout = QVBoxLayout(form_widget)
        form_layout.setContentsMargins(5, 5, 5, 5)
        form_layout.setSpacing(15)

        # Поля данных товара
        fields = [
            ("Название товара:", "product_name_input", QLineEdit()),
            ("Артикул:", "product_article_input", QLineEdit()),
            ("Тип:", "product_type_input", QLineEdit()),
            ("Материал:", "product_material_input", QLineEdit()),
            ("Тип вставки:", "product_insert_input", QLineEdit()),
            ("Вес (г):", "product_weight_input", QLineEdit()),
            ("Цена (₽):", "product_price_input", QLineEdit()),
            ("Количество на складе:", "product_quantity_input", QLineEdit())
        ]

        for label_text, attr_name, widget in fields:
            setattr(self, attr_name, widget)
            field_layout = QHBoxLayout()
            field_layout.addWidget(QLabel(label_text), stretch=1)
            field_layout.addWidget(widget, stretch=2)
            form_layout.addLayout(field_layout)

        # Группа для загрузки изображения
        image_group = QGroupBox("Изображение товара")
        image_layout = QVBoxLayout(image_group)
        image_layout.setSpacing(15)

        # Кнопка выбора файла
        browse_btn = QPushButton("Выбрать изображение")
        browse_btn.setObjectName("browseButton")
        browse_btn.clicked.connect(self.select_image_file)
        image_layout.addWidget(browse_btn, alignment=Qt.AlignmentFlag.AlignLeft)

        # Область превью изображения
        self.image_preview = QLabel()
        self.image_preview.setMinimumSize(300, 200)
        self.image_preview.setStyleSheet("""
            QLabel {
                background-color: #333;
                border: 1px dashed #555;
                qproperty-alignment: 'AlignCenter';
            }
        """)
        image_layout.addWidget(self.image_preview)

        # Статус выбранного файла
        self.image_path_label = QLabel("Файл не выбран")
        self.image_path_label.setStyleSheet("color: #888; font-size: 12px;")
        image_layout.addWidget(self.image_path_label, alignment=Qt.AlignmentFlag.AlignCenter)

        form_layout.addWidget(image_group)
        form_scroll.setWidget(form_widget)
        main_layout.addWidget(form_scroll)

        # Кнопки управления внизу
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setObjectName("cancelButton")
        cancel_btn.clicked.connect(dialog.reject)

        save_btn = QPushButton("Сохранить")
        save_btn.setObjectName("saveButton")
        save_btn.clicked.connect(lambda: self.save_new_product_with_image(dialog))

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(save_btn)
        main_layout.addLayout(buttons_layout)

        # Статусная строка
        self.add_product_status = QLabel()
        self.add_product_status.setWordWrap(True)
        self.add_product_status.setStyleSheet("""
            QLabel {
                color: #ff5555;
                font-size: 13px;
                margin-top: 10px;
                padding: 5px;
            }
        """)
        main_layout.addWidget(self.add_product_status)

        self.current_image_path = None
        dialog.exec()

    def select_image_file(self, update_preview=False):
        file_dialog = QFileDialog()
        file_dialog.setStyleSheet("""
            QFileDialog {
                background-color: #1a1a1a;
                color: white;
            }
            QLabel {
                color: white;
            }
            QPushButton {
                background-color: #444;
                color: white;
                padding: 5px 10px;
                border-radius: 4px;
            }
        """)
        file_dialog.setNameFilter("Images (*.png *.jpg *.jpeg *.bmp)")
        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            if selected_files:
                self.current_image_path = selected_files[0]
                self.image_path_label.setText(os.path.basename(self.current_image_path))
                self.image_path_label.setStyleSheet("color: #d4af37; font-size: 12px;")

                pixmap = QPixmap(self.current_image_path)
                if not pixmap.isNull():
                    pixmap = pixmap.scaled(
                        300, 200,
                        Qt.AspectRatioMode.KeepAspectRatio,
                        Qt.TransformationMode.SmoothTransformation
                    )
                    # Обновляем превью в зависимости от контекста
                    if update_preview and hasattr(self, 'edit_image_preview'):
                        self.edit_image_preview.setPixmap(pixmap)
                        self.edit_image_status.setText("Новое изображение (будет сохранено)")
                    else:
                        self.image_preview.setPixmap(pixmap)

    def save_new_product_with_image(self, dialog):
        self.log_action("Создание нового товара")
        if not all([
            self.product_name_input.text().strip(),
            self.product_article_input.text().strip(),
            self.product_type_input.text().strip(),
            self.product_material_input.text().strip(),
            self.product_weight_input.text().strip(),
            self.product_price_input.text().strip(),
            self.product_quantity_input.text().strip()
        ]):
            self.log_action("Ошибка: не заполнены обязательные поля")
            self.add_product_status.setText("Заполните все обязательные поля")
            self.add_product_status.setStyleSheet("color: #ff5555;")
            return

        try:
            image_id = 0
            if self.current_image_path:
                with open(self.current_image_path, 'rb') as img_file:
                    files = {'file': (self.current_image_path.split('/')[-1], img_file, 'image/jpeg')}
                    self.log_action("Загрузка изображения товара")
                    response = requests.post(
                        f"{self.base_url}/images/upload/",
                        files=files
                    )
                    if response.status_code == 200:
                        image_id = response.json().get('id', 0)
                        self.log_action(f"Изображение загружено, ID: {image_id}")
                    else:
                        raise Exception(f"Ошибка загрузки изображения: {response.text}")

            product_data = {
                "name": self.product_name_input.text().strip(),
                "article": self.product_article_input.text().strip(),
                "type": self.product_type_input.text().strip(),
                "material": self.product_material_input.text().strip(),
                "insert_type": self.product_insert_input.text().strip(),
                "weight": float(self.product_weight_input.text()),
                "price": float(self.product_price_input.text()),
                "stock_quantity": int(self.product_quantity_input.text()),
                "image_id": image_id
            }

            self.log_action("Сохранение данных товара")
            response = requests.post(
                f"{self.base_url}/products/create",
                json=product_data,
                headers={"Content-Type": "application/json"}
            )

            if response.status_code == 200:
                self.log_action(f"Товар успешно создан, ID: {response.json().get('product_id')}")
                self.add_product_status.setText("Товар успешно добавлен")
                self.add_product_status.setStyleSheet("color: #55ff55;")
                QTimer.singleShot(1500, dialog.close)
                self.load_products()
            else:
                self.log_action(f"Ошибка создания товара: {response.text}")
                self.add_product_status.setText(f"Ошибка при добавлении товара: {response.text}")
                self.add_product_status.setStyleSheet("color: #ff5555;")

        except Exception as e:
            self.log_action(f"Ошибка при создании товара: {str(e)}")
            self.add_product_status.setText(f"Ошибка: {str(e)}")
            self.add_product_status.setStyleSheet("color: #ff5555;")

    def save_new_product(self, dialog):
        """Сохраняет новый товар"""
        # Получаем данные из формы
        product_data = {
            "name": self.product_name_input.text().strip(),
            "article": self.product_article_input.text().strip(),
            "type": self.product_type_combo.currentText(),
            "material": self.product_material_combo.currentText(),
            "insert_type": self.product_insert_combo.currentText(),
            "weight": float(self.product_weight_input.text()) if self.product_weight_input.text() else 0,
            "price": float(self.product_price_input.text()) if self.product_price_input.text() else 0,
            "stock_quantity": int(self.product_quantity_input.text()) if self.product_quantity_input.text() else 0,
            "image_id": int(self.product_image_id_input.text()) if self.product_image_id_input.text() else 0
        }

        # Проверяем обязательные поля
        if not product_data["name"] or not product_data["article"]:
            self.add_product_status.setText("Название и артикул обязательны для заполнения")
            self.add_product_status.setStyleSheet("color: #ff5555;")
            return

        try:
            response = requests.post(
                f"{self.base_url}/products/create",
                json=product_data,
                headers={"Content-Type": "application/json"}
            )

            if response.status_code == 200:
                self.add_product_status.setText("Товар успешно добавлен")
                self.add_product_status.setStyleSheet("color: #55ff55;")
                QTimer.singleShot(1500, dialog.close)
                self.load_products()  # Обновляем список товаров
            else:
                self.add_product_status.setText(f"Ошибка при добавлении товара: {response.text}")
                self.add_product_status.setStyleSheet("color: #ff5555;")
        except Exception as e:
            self.add_product_status.setText(f"Ошибка: {str(e)}")
            self.add_product_status.setStyleSheet("color: #ff5555;")

    def search_products(self):
        search_text = self.product_search.text().strip().lower()
        self.log_action(f"Поиск товаров: '{search_text}'")

        try:
            response = requests.get(f"{self.base_url}/products/get/all/")
            if response.status_code == 200:
                products = response.json()
                filtered_products = []

                for product in products:
                    matches = [
                        search_text in product.get('name', '').lower(),
                        search_text in product.get('article', '').lower()
                    ]

                    if any(matches):
                        filtered_products.append(product)

                self.populate_products_table(filtered_products)
                self.log_action(f"Найдено {len(filtered_products)} товаров")
            else:
                self.log_action(f"Ошибка поиска товаров: код {response.status_code}")
                QMessageBox.warning(self, "Ошибка", f"Не удалось выполнить поиск. Код ошибки: {response.status_code}")
        except Exception as e:
            self.log_action(f"Ошибка при поиске товаров: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при поиске товаров: {str(e)}")

    def update_product_buttons_state(self):
        """Обновляет состояние кнопок в зависимости от выбранной строки"""
        selected = self.products_table.selectionModel().hasSelection()
        self.edit_product_btn.setEnabled(selected)
        self.delete_product_btn.setEnabled(selected)

    def edit_selected_product(self):
        """Редактирование выбранного товара"""
        selected_row = self.products_table.currentRow()
        if selected_row >= 0:
            product_id = int(self.products_table.item(selected_row, 0).text())
            product_data = {
                "product_id": product_id,
                "name": self.products_table.item(selected_row, 1).text(),
                "article": self.products_table.item(selected_row, 2).text(),
                "type": self.products_table.item(selected_row, 3).text(),
                "material": self.products_table.item(selected_row, 4).text(),
                "insert_type": self.products_table.item(selected_row, 5).text(),
                "weight": float(self.products_table.item(selected_row, 6).text()),
                "price": float(self.products_table.item(selected_row, 7).text().replace(" ₽", "").replace(" ", "")),
                "stock_quantity": int(self.products_table.item(selected_row, 8).text()),
                "image_id": int(self.products_table.item(selected_row, 9).text()) if self.products_table.item(
                    selected_row, 9) else 0
            }
            self.show_edit_product_dialog(product_data)

    def show_edit_product_dialog(self, product_data):
        """Диалог редактирования товара"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Редактирование товара ID: {product_data['product_id']}")
        dialog.setFixedSize(600, 800)

        # Основной стиль диалога
        dialog.setStyleSheet("""
            QDialog {
                background-color: #1a1a1a;
                color: white;
                font-family: 'Segoe UI';
            }
            QLabel {
                color: white;
            }
            QLineEdit, QTextEdit {
                background-color: #333;
                color: white;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
                selection-background-color: #d4af37;
                selection-color: black;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 14px;
                min-width: 100px;
            }
            #saveButton {
                background-color: #d4af37;
                color: black;
            }
            #saveButton:hover {
                background-color: #f0d87c;
            }
            #cancelButton {
                background-color: #444;
                color: white;
            }
            #cancelButton:hover {
                background-color: #555;
            }
            #browseButton {
                background-color: #555;
                color: white;
            }
            #browseButton:hover {
                background-color: #666;
            }
            QGroupBox {
                color: #d4af37;
                border: 1px solid #444;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)

        # Сохраняем оригинальные данные
        self.original_product_data = product_data.copy()
        self.current_image_path = None
        self.new_image_id = None

        # Основной layout
        main_layout = QVBoxLayout(dialog)
        main_layout.setContentsMargins(25, 25, 25, 25)
        main_layout.setSpacing(15)

        # Заголовок
        title = QLabel(f"Редактирование товара: {product_data['name']}")
        title.setFont(QFont('Montserrat', 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #d4af37; margin-bottom: 15px;")
        main_layout.addWidget(title)

        # Форма с полями
        form_scroll = QScrollArea()
        form_scroll.setWidgetResizable(True)
        form_scroll.setStyleSheet("border: none;")

        form_widget = QWidget()
        form_layout = QVBoxLayout(form_widget)
        form_layout.setContentsMargins(5, 5, 5, 5)
        form_layout.setSpacing(15)

        # Поля для редактирования
        fields = [
            ("Название:", "edit_name_input", QLineEdit(product_data["name"])),
            ("Артикул:", "edit_article_input", QLineEdit(product_data["article"])),
            ("Тип:", "edit_type_input", QLineEdit(product_data["type"])),
            ("Материал:", "edit_material_input", QLineEdit(product_data["material"])),
            ("Тип вставки:", "edit_insert_input", QLineEdit(product_data["insert_type"])),
            ("Вес (г):", "edit_weight_input", QLineEdit(str(product_data["weight"]))),
            ("Цена (₽):", "edit_price_input", QLineEdit(str(product_data["price"]))),
            ("Количество:", "edit_quantity_input", QLineEdit(str(product_data["stock_quantity"])))
        ]

        for label_text, attr_name, widget in fields:
            setattr(self, attr_name, widget)
            field_layout = QHBoxLayout()
            field_layout.addWidget(QLabel(label_text), stretch=1)
            field_layout.addWidget(widget, stretch=2)
            form_layout.addLayout(field_layout)

        # Группа изображения
        image_group = QGroupBox("Изображение товара")
        image_layout = QVBoxLayout(image_group)
        image_layout.setSpacing(15)

        # Элементы для работы с изображением
        self.edit_image_path_label = QLabel("Файл не выбран")
        self.edit_image_path_label.setStyleSheet("color: #888; font-size: 12px;")

        self.edit_image_preview = QLabel()
        self.edit_image_preview.setMinimumSize(300, 200)
        self.edit_image_preview.setStyleSheet("""
            QLabel {
                background-color: #333;
                border: 1px dashed #555;
                qproperty-alignment: 'AlignCenter';
            }
        """)

        # Загружаем текущее изображение если есть
        if product_data["image_id"]:
            self.load_product_image_preview(product_data["image_id"], self.edit_image_preview)
            self.edit_image_status = QLabel(f"Текущее изображение (ID: {product_data['image_id']})")
        else:
            self.edit_image_status = QLabel("Изображение отсутствует")

        self.edit_image_status.setStyleSheet("color: #888; font-size: 12px;")

        browse_btn = QPushButton("Изменить изображение")
        browse_btn.setObjectName("browseButton")
        browse_btn.clicked.connect(self.select_edit_image_file)

        image_layout.addWidget(browse_btn, alignment=Qt.AlignmentFlag.AlignLeft)
        image_layout.addWidget(self.edit_image_preview)
        image_layout.addWidget(self.edit_image_path_label, alignment=Qt.AlignmentFlag.AlignCenter)
        image_layout.addWidget(self.edit_image_status, alignment=Qt.AlignmentFlag.AlignCenter)

        form_layout.addWidget(image_group)
        form_scroll.setWidget(form_widget)
        main_layout.addWidget(form_scroll)

        # Кнопки управления
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setObjectName("cancelButton")
        cancel_btn.clicked.connect(dialog.reject)

        save_btn = QPushButton("Сохранить изменения")
        save_btn.setObjectName("saveButton")
        save_btn.clicked.connect(lambda: self.update_product(product_data["product_id"], dialog))

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(save_btn)
        main_layout.addLayout(buttons_layout)

        # Статусная строка
        self.edit_status_label = QLabel()
        self.edit_status_label.setWordWrap(True)
        self.edit_status_label.setStyleSheet("""
            QLabel {
                color: #ff5555;
                font-size: 13px;
                margin-top: 10px;
                padding: 5px;
            }
        """)
        main_layout.addWidget(self.edit_status_label)

        dialog.exec()

    def select_edit_image_file(self):
        """Выбор изображения при редактировании товара"""
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Images (*.png *.jpg *.jpeg *.bmp)")
        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            if selected_files:
                self.current_image_path = selected_files[0]
                self.edit_image_path_label.setText(os.path.basename(self.current_image_path))
                self.edit_image_path_label.setStyleSheet("color: #d4af37;")

                pixmap = QPixmap(self.current_image_path)
                if not pixmap.isNull():
                    pixmap = pixmap.scaled(300, 200, Qt.AspectRatioMode.KeepAspectRatio)
                    self.edit_image_preview.setPixmap(pixmap)
                    self.edit_image_status.setText("Новое изображение (не сохранено)")

    def update_product(self, product_id, dialog):
        self.log_action(f"Обновление товара ID: {product_id}")
        try:
            update_data = {
                "name": self.edit_name_input.text().strip(),
                "article": self.edit_article_input.text().strip(),
                "type": self.edit_type_input.text().strip(),
                "material": self.edit_material_input.text().strip(),
                "insert_type": self.edit_insert_input.text().strip(),
                "weight": float(self.edit_weight_input.text()),
                "price": float(self.edit_price_input.text()),
                "stock_quantity": int(self.edit_quantity_input.text()),
                "image_id": self.original_product_data["image_id"]
            }

            if self.current_image_path:
                with open(self.current_image_path, 'rb') as img_file:
                    files = {'file': (os.path.basename(self.current_image_path), img_file, 'image/jpeg')}
                    upload_response = requests.post(
                        f"{self.base_url}/images/upload/",
                        files=files
                    )

                    if upload_response.status_code != 200:
                        raise Exception(f"Ошибка загрузки изображения: {upload_response.text}")

                    new_image_id = upload_response.json().get('id')
                    if not new_image_id:
                        raise Exception("Не получен ID нового изображения")

                    update_data["image_id"] = new_image_id
                    self.log_action(f"Загружено новое изображение для товара ID: {product_id}")

            update_response = requests.patch(
                f"{self.base_url}/products/update/{product_id}",
                json=update_data,
                headers={"Content-Type": "application/json"}
            )

            if update_response.status_code == 200:
                self.log_action(f"Товар ID: {product_id} успешно обновлен")
                self.edit_status_label.setText("Товар успешно обновлен")
                self.edit_status_label.setStyleSheet("color: #55ff55;")
                QTimer.singleShot(1500, lambda: (dialog.close(), self.load_products()))
            else:
                error_msg = update_response.json().get('detail', update_response.text)
                self.log_action(f"Ошибка обновления товара ID: {product_id}: {error_msg}")
                raise Exception(f"Ошибка обновления товара: {error_msg}")

        except Exception as e:
            self.log_action(f"Ошибка при обновлении товара ID: {product_id}: {str(e)}")
            self.edit_status_label.setText(f"Ошибка: {str(e)}")
            self.edit_status_label.setStyleSheet("color: #ff5555;")
            if hasattr(self, 'original_product_data') and self.original_product_data.get("image_id"):
                self.load_product_image_preview(self.original_product_data["image_id"], self.edit_image_preview)
                self.edit_image_status.setText(f"Текущее изображение (ID: {self.original_product_data['image_id']})")

    def load_product_image_preview(self, image_id, preview_label):
        """Загружает изображение товара для предпросмотра"""
        try:
            response = requests.get(f"{self.base_url}/images/{image_id}")
            if response.status_code == 200:
                pixmap = QPixmap()
                pixmap.loadFromData(response.content)
                if not pixmap.isNull():
                    pixmap = pixmap.scaled(300, 200, Qt.AspectRatioMode.KeepAspectRatio)
                    preview_label.setPixmap(pixmap)
        except Exception as e:
            print(f"Ошибка загрузки изображения: {str(e)}")

    def delete_selected_product(self):
        """Удаление выбранного товара"""
        selected_row = self.products_table.currentRow()
        if selected_row >= 0:
            product_id = int(self.products_table.item(selected_row, 0).text())
            product_name = self.products_table.item(selected_row, 1).text()

            # Диалог подтверждения
            msg_box = QMessageBox()
            msg_box.setWindowTitle("Подтверждение удаления")
            msg_box.setText(f"Вы уверены, что хотите удалить товар:\n{product_name} (ID: {product_id})?")
            msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            msg_box.setStyleSheet("""
                QMessageBox {
                    background-color: #1a1a1a;
                }
                QLabel {
                    color: white;
                }
                QPushButton {
                    color: white;
                    background-color: #333;
                    padding: 5px 10px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #444;
                }
            """)
            msg_box.button(QMessageBox.StandardButton.Yes).setText("Да")
            msg_box.button(QMessageBox.StandardButton.No).setText("Нет")

            if msg_box.exec() == QMessageBox.StandardButton.Yes:
                try:
                    response = requests.delete(f"{self.base_url}/products/delete/{product_id}")
                    if response.status_code == 200:
                        self.load_products()
                    else:
                        QMessageBox.warning(self, "Ошибка", f"Не удалось удалить товар: {response.text}")
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Ошибка соединения: {str(e)}")

    def delete_product(self, product_id):
        msg_box = QMessageBox()
        msg_box.setWindowTitle("Подтверждение удаления")
        msg_box.setText(f"Вы уверены, что хотите удалить товар с ID {product_id}?")
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #1a1a1a;
            }
            QLabel {
                color: white;
            }
            QPushButton {
                color: white;
                background-color: #333;
                padding: 5px 10px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #444;
            }
        """)
        msg_box.button(QMessageBox.StandardButton.Yes).setText("Да")
        msg_box.button(QMessageBox.StandardButton.No).setText("Нет")

        reply = msg_box.exec()

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.log_action(f"Удаление товара ID: {product_id}")
                response = requests.delete(f"{self.base_url}/products/{product_id}")

                if response.status_code == 200:
                    self.log_action(f"Товар ID: {product_id} успешно удален")
                    QMessageBox.information(self, "Успех", "Товар успешно удален")
                    self.load_products()
                else:
                    self.log_action(f"Ошибка удаления товара ID: {product_id}: {response.text}")
                    QMessageBox.warning(self, "Ошибка", f"Не удалось удалить товар: {response.text}")
            except Exception as e:
                self.log_action(f"Ошибка при удалении товара ID: {product_id}: {str(e)}")
                QMessageBox.critical(self, "Ошибка", f"Ошибка соединения: {str(e)}")

    def load_products(self):
        try:
            self.log_action("Запрос списка товаров")
            response = requests.get(f"{self.base_url}/products/get/all/")

            if response.status_code == 200:
                products = response.json()
                self.populate_products_table(products)
                self.update_dashboard_stats(None, len(products))
                self.log_action(f"Успешно загружено {len(products)} товаров")
            else:
                self.log_action(f"Ошибка загрузки товаров: код {response.status_code}")
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить товары. Код ошибки: {response.status_code}")
        except Exception as e:
            self.log_action(f"Ошибка при загрузке товаров: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке товаров: {str(e)}")

    def populate_products_table(self, products):
        self.products_table.setRowCount(0)

        for product in products:
            row_position = self.products_table.rowCount()
            self.products_table.insertRow(row_position)

            # Заполняем основные данные
            self.products_table.setItem(row_position, 0, QTableWidgetItem(str(product.get('product_id', ''))))
            self.products_table.setItem(row_position, 1, QTableWidgetItem(product.get('name', '')))
            self.products_table.setItem(row_position, 2, QTableWidgetItem(product.get('article', '')))
            self.products_table.setItem(row_position, 3, QTableWidgetItem(product.get('type', '')))
            self.products_table.setItem(row_position, 4, QTableWidgetItem(product.get('material', '')))
            self.products_table.setItem(row_position, 5, QTableWidgetItem(product.get('insert_type', '')))
            self.products_table.setItem(row_position, 6, QTableWidgetItem(str(product.get('weight', 0))))
            self.products_table.setItem(row_position, 7,
                                        QTableWidgetItem(f"{product.get('price', 0):,} ₽".replace(",", " ")))
            self.products_table.setItem(row_position, 8, QTableWidgetItem(str(product.get('stock_quantity', 0))))
            self.products_table.setItem(row_position, 9, QTableWidgetItem(str(product.get('image_id', 0))))

            # Обработка даты создания
            created_at = product.get('created_at')
            date_item = QTableWidgetItem()

            if created_at:
                try:
                    # Универсальный парсинг даты
                    if '.' in created_at:
                        dt = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%S.%f")
                    else:
                        dt = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%S")

                    formatted_date = dt.strftime("%d.%m.%Y %H:%M")
                    date_item.setText(formatted_date)
                except Exception as e:
                    print(f"Ошибка форматирования даты {created_at}: {str(e)}")
                    date_item.setText(created_at)
            else:
                date_item.setText("Нет данных")

            self.products_table.setItem(row_position, 10, date_item)

            image_id = product.get('image_id', 0)
            if image_id:
                btn = QPushButton("Просмотр")
                btn.setObjectName("viewImageButton")
                btn.setStyleSheet("""
                            QPushButton {
                                background-color: #d4af37;
                                color: black;
                                padding: 5px;
                                border-radius: 4px;
                            }
                            QPushButton:hover {
                                background-color: #f0d87c;
                            }
                        """)
                btn.clicked.connect(lambda _, img_id=image_id: self.view_product_image(img_id))
                self.products_table.setCellWidget(row_position, 11, btn)
            else:
                self.products_table.setItem(row_position, 11, QTableWidgetItem("Нет изображения"))

    def view_product_image(self, image_id):
        """Открывает окно с изображением товара"""
        try:
            # Загружаем изображение с сервера
            response = requests.get(f"{self.base_url}/images/{image_id}")

            if response.status_code == 200:
                # Создаем диалоговое окно для отображения изображения
                dialog = QDialog(self)
                dialog.setWindowTitle(f"Изображение товара (ID: {image_id})")
                dialog.setMinimumSize(400, 400)

                layout = QVBoxLayout(dialog)

                # Создаем QLabel для отображения изображения
                image_label = QLabel()
                image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

                # Загружаем изображение из ответа сервера
                image_data = response.content
                pixmap = QPixmap()
                pixmap.loadFromData(image_data)

                # Масштабируем изображение, если оно слишком большое
                if pixmap.width() > 800 or pixmap.height() > 600:
                    pixmap = pixmap.scaled(800, 600, Qt.AspectRatioMode.KeepAspectRatio)

                image_label.setPixmap(pixmap)

                # Кнопка закрытия
                close_btn = QPushButton("Закрыть")
                close_btn.clicked.connect(dialog.close)

                layout.addWidget(image_label)
                layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)

                dialog.exec()
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить изображение. Код ошибки: {response.status_code}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке изображения: {str(e)}")

    def create_orders_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        title = QLabel("Управление заказами")
        title.setFont(QFont('Montserrat', 20, QFont.Weight.Bold))
        title.setStyleSheet("color: gold;")
        layout.addWidget(title)

        # Filters
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.status_combo = QComboBox()
        self.status_combo.addItems(["Все статусы", "В обработке", "Завершен", "Выполнено", "Отменен"])

        # Date range filters
        self.date_from_edit = QDateEdit()
        self.date_from_edit.setCalendarPopup(True)
        self.date_from_edit.setDate(QDate.currentDate().addMonths(-1))

        self.date_to_edit = QDateEdit()
        self.date_to_edit.setCalendarPopup(True)
        self.date_to_edit.setDate(QDate.currentDate())

        filter_layout.addWidget(self.status_combo)
        filter_layout.addWidget(QLabel("с"))
        filter_layout.addWidget(self.date_from_edit)
        filter_layout.addWidget(QLabel("до"))
        filter_layout.addWidget(self.date_to_edit)
        filter_layout.addStretch()

        apply_filters_btn = QPushButton("Применить фильтры")
        apply_filters_btn.setObjectName("applyButton")
        apply_filters_btn.clicked.connect(self.apply_order_filters)
        filter_layout.addWidget(apply_filters_btn)

        layout.addLayout(filter_layout)

        # Orders table
        self.orders_table = QTableWidget()
        self.orders_table.setColumnCount(4)  # Уменьшаем количество колонок до 5
        self.orders_table.setHorizontalHeaderLabels(["ID", "Клиент", "Дата", "Статус"])
        self.orders_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.orders_table.verticalHeader().setVisible(False)
        self.orders_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.orders_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        # Загружаем заказы при открытии страницы
        self.load_orders()

        layout.addWidget(self.orders_table)

        # Order details button
        details_btn = QPushButton("Просмотреть детали")
        details_btn.setObjectName("detailsButton")
        details_btn.clicked.connect(self.show_order_details_admin)
        layout.addWidget(details_btn, alignment=Qt.AlignmentFlag.AlignRight)

        return page

    def apply_order_filters(self):
        """Применяет выбранные фильтры к заказам"""
        status_filter = self.status_combo.currentText()

        # Получаем даты из QDateEdit
        date_from_qdate = self.date_from_edit.date()
        date_to_qdate = self.date_to_edit.date()

        # Преобразуем QDate в date
        date_from = date(date_from_qdate.year(), date_from_qdate.month(), date_from_qdate.day())
        date_to = date(date_to_qdate.year(), date_to_qdate.month(), date_to_qdate.day())

        filtered_orders = []
        for order in getattr(self, 'all_orders', []):
            # Проверяем статус
            status_match = (status_filter == "Все статусы") or (order.get('status', '') == status_filter)

            # Проверяем дату
            order_date_str = order.get('order_date', '')
            if order_date_str:
                try:
                    # Преобразуем строку даты в объект даты
                    order_date = datetime.fromisoformat(order_date_str.replace('Z', '+00:00')).date()
                    date_match = date_from <= order_date <= date_to
                except ValueError:
                    date_match = False
            else:
                date_match = False

            if status_match and date_match:
                filtered_orders.append(order)

        self.populate_orders_table(filtered_orders)

    def load_orders(self):
        try:
            self.log_action("Запрос списка заказов")
            response = requests.get(f"{self.base_url}/orders/get/all")

            if response.status_code == 200:
                orders = response.json()
                self.all_orders = orders
                self.populate_orders_table(orders)
                self.update_dashboard_stats(None, None, len(orders))
                self.log_action(f"Успешно загружено {len(orders)} заказов")
            else:
                self.log_action(f"Ошибка загрузки заказов: код {response.status_code}")
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить заказы. Код ошибки: {response.status_code}")
        except Exception as e:
            self.log_action(f"Ошибка при загрузке заказов: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке заказов: {str(e)}")

    def populate_orders_table(self, orders):
        self.orders_table.setRowCount(0)

        for order in orders:
            row_position = self.orders_table.rowCount()
            self.orders_table.insertRow(row_position)

            # Заполняем данные заказа
            self.orders_table.setItem(row_position, 0, QTableWidgetItem(str(order.get('order_id', ''))))
            self.orders_table.setItem(row_position, 1, QTableWidgetItem(order.get('username', '')))

            # Форматируем дату
            order_date = order.get('order_date', '')
            if order_date:
                try:
                    dt = datetime.fromisoformat(order_date.replace('Z', '+00:00'))
                    date_str = dt.strftime("%d.%m.%Y %H:%M")
                except ValueError:
                    date_str = order_date
            else:
                date_str = "Нет данных"
            self.orders_table.setItem(row_position, 2, QTableWidgetItem(date_str))

            self.orders_table.setItem(row_position, 3, QTableWidgetItem(order.get('status', '')))

    def create_returns_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        title = QLabel("Управление возвратами")
        title.setFont(QFont('Montserrat', 20, QFont.Weight.Bold))
        title.setStyleSheet("color: gold;")
        layout.addWidget(title)

        # Filters
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.return_status_combo = QComboBox()
        self.return_status_combo.addItems(["Все статусы", "В обработке", "Одобрен", "Отклонен", "Завершен"])

        self.return_address_combo = QComboBox()
        self.return_address_combo.addItem("Все адреса")

        # Кнопка обновления адресов
        refresh_address_btn = QPushButton("Обновить адреса")
        refresh_address_btn.setObjectName("refreshButton")
        refresh_address_btn.clicked.connect(self.load_return_addresses)

        filter_layout.addWidget(self.return_status_combo)
        filter_layout.addWidget(self.return_address_combo)
        filter_layout.addWidget(refresh_address_btn)
        filter_layout.addStretch()

        apply_filters_btn = QPushButton("Применить фильтры")
        apply_filters_btn.setObjectName("applyButton")
        apply_filters_btn.clicked.connect(self.apply_return_filters)
        filter_layout.addWidget(apply_filters_btn)

        layout.addLayout(filter_layout)

        # Returns table
        self.returns_table = QTableWidget()
        self.returns_table.setColumnCount(7)
        self.returns_table.setHorizontalHeaderLabels(
            ["ID возврата", "ID заказа", "Клиент", "Дата", "Статус", "Адрес", "Причина"])
        self.returns_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.returns_table.verticalHeader().setVisible(False)

        # Загружаем возвраты при открытии страницы
        self.load_returns()

        layout.addWidget(self.returns_table)

        return page

    def apply_return_filters(self):
        """Применяет выбранные фильтры к возвратам"""
        status_filter = self.return_status_combo.currentText()
        address_filter = self.return_address_combo.currentText()

        filtered_returns = []
        for return_item in getattr(self, 'all_returns', []):
            # Проверяем статус
            status_match = (status_filter == "Все статусы") or (return_item.get('status', '') == status_filter)

            # Проверяем адрес (нужно получать из заказа)
            address_match = (
                        address_filter == "Все адреса")  # or (get_order_address(return_item['order_id']) == address_filter)

            if status_match and address_match:
                filtered_returns.append(return_item)

        self.populate_returns_table(filtered_returns)

    def load_returns(self):
        try:
            self.log_action("Запрос списка возвратов")
            response = requests.get(f"{self.base_url}/returns/get/all/")

            if response.status_code == 200:
                returns = response.json()
                self.all_returns = returns
                self.populate_returns_table(returns)
                self.update_dashboard_stats(None, None, None, len(returns))
                self.log_action(f"Успешно загружено {len(returns)} возвратов")
            else:
                self.log_action(f"Ошибка загрузки возвратов: код {response.status_code}")
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить возвраты. Код ошибки: {response.status_code}")
        except Exception as e:
            self.log_action(f"Ошибка при загрузке возвратов: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке возвратов: {str(e)}")

    def populate_returns_table(self, returns):
        self.returns_table.setRowCount(0)

        for return_item in returns:
            row_position = self.returns_table.rowCount()
            self.returns_table.insertRow(row_position)

            # Заполняем данные возврата
            self.returns_table.setItem(row_position, 0, QTableWidgetItem(str(return_item.get('return_id', ''))))
            self.returns_table.setItem(row_position, 1, QTableWidgetItem(str(return_item.get('order_id', ''))))

            # Клиента нужно получать из заказа, так как в возврате только client_id
            client_id = return_item.get('client_id', '')
            self.returns_table.setItem(row_position, 2, QTableWidgetItem(f"Клиент {client_id}"))

            # Форматируем дату
            return_date = return_item.get('return_date', '')
            if return_date:
                try:
                    dt = datetime.fromisoformat(return_date.replace('Z', '+00:00'))
                    date_str = dt.strftime("%d.%m.%Y %H:%M")
                except ValueError:
                    date_str = return_date
            else:
                date_str = "Нет данных"
            self.returns_table.setItem(row_position, 3, QTableWidgetItem(date_str))

            self.returns_table.setItem(row_position, 4, QTableWidgetItem(return_item.get('status', '')))

            # Адрес нужно получать из заказа - временная заглушка
            self.returns_table.setItem(row_position, 5, QTableWidgetItem("Адрес не указан"))

            self.returns_table.setItem(row_position, 6, QTableWidgetItem(return_item.get('description', '')))

    def load_return_addresses(self):
        """Загружает уникальные адреса из возвратов для фильтрации"""
        try:
            # В реальном API нужно получать адреса из связанных заказов
            # Здесь временная реализация
            addresses = set()
            for return_item in getattr(self, 'all_returns', []):
                # В вашем API нужно получать адрес из заказа по order_id
                # addresses.add(get_order_address(return_item['order_id']))
                pass

            # Добавляем тестовые данные
            addresses.update(["ул. Ленина, д.10", "пр. Мира, д.5", "ул. Пушкина, д.15"])

            self.return_address_combo.clear()
            self.return_address_combo.addItem("Все адреса")
            for address in sorted(addresses):
                if address:  # Пропускаем пустые адреса
                    self.return_address_combo.addItem(address)

        except Exception as e:
            print(f"Ошибка загрузки адресов: {str(e)}")

    def create_reports_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        title = QLabel("Отчеты")
        title.setFont(QFont('Montserrat', 20, QFont.Weight.Bold))
        title.setStyleSheet("color: gold;")
        layout.addWidget(title)

        # Report type selection
        report_group = QGroupBox("Тип отчета")
        report_group.setStyleSheet("""
            QGroupBox {
                color: #d4af37;
                border: 1px solid #444;
                border-radius: 5px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)

        report_type_layout = QHBoxLayout(report_group)

        self.report_type_combo = QComboBox()
        self.report_type_combo.addItems([
            "Отчет по продажам",
            "Отчет по возвратам",
            "Отчет по товарам",
            "Отчет по клиентам",
            "Сводный отчет"
        ])
        report_type_layout.addWidget(self.report_type_combo)
        report_type_layout.addStretch()

        layout.addWidget(report_group)

        # Date range filter
        filter_group = QGroupBox("Параметры отчета")
        filter_group.setStyleSheet(report_group.styleSheet())

        filter_layout = QHBoxLayout(filter_group)
        filter_layout.setSpacing(10)

        # Date range
        filter_layout.addWidget(QLabel("Период:"))

        self.report_date_from = QDateEdit()
        self.report_date_from.setCalendarPopup(True)
        self.report_date_from.setDate(QDate.currentDate().addMonths(-1))

        self.report_date_to = QDateEdit()
        self.report_date_to.setCalendarPopup(True)
        self.report_date_to.setDate(QDate.currentDate())

        filter_layout.addWidget(self.report_date_from)
        filter_layout.addWidget(QLabel("до"))
        filter_layout.addWidget(self.report_date_to)
        filter_layout.addStretch()

        layout.addWidget(filter_group)

        # Buttons
        button_layout = QHBoxLayout()

        generate_btn = QPushButton("Сформировать отчет")
        generate_btn.setObjectName("generateButton")
        generate_btn.clicked.connect(self.generate_report_from_data)
        button_layout.addWidget(generate_btn)

        export_btn = QPushButton("Экспорт в Excel")
        export_btn.setObjectName("exportButton")
        export_btn.clicked.connect(self.export_report_to_xlsx)  # Изменено на export_report_to_xlsx
        button_layout.addWidget(export_btn)

        layout.addLayout(button_layout)

        # Report preview area
        self.report_preview = QTextEdit()
        self.report_preview.setReadOnly(True)
        self.report_preview.setStyleSheet("""
            QTextEdit {
                background-color: #252525;
                border: 1px solid #444;
                border-radius: 5px;
                color: white;
                font-family: 'Consolas';
                font-size: 14px;
            }
        """)
        self.report_preview.setPlainText(
            "Здесь будет отображаться сформированный отчет.\n"
            "Выберите тип отчета и период, затем нажмите 'Сформировать отчет'."
        )

        layout.addWidget(self.report_preview, stretch=1)

        return page

    def export_report_to_xlsx(self):
        """Экспортирует текущий отчет в XLSX файл (Excel)"""
        report_text = self.report_preview.toPlainText()
        if not report_text or "Здесь будет отображаться" in report_text:
            QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта. Сначала сформируйте отчет.")
            return

        options = QFileDialog.Option(0)
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт отчета в Excel",
            "",
            "Excel Files (*.xlsx)",
            options=options
        )

        if file_name:
            try:
                if not file_name.endswith('.xlsx'):
                    file_name += '.xlsx'

                # Создаем новую книгу Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Отчет"

                # Настройка стилей
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                # Разбираем текстовый отчет и переносим в Excel
                lines = report_text.split('\n')
                current_row = 1

                for line in lines:
                    if not line.strip():
                        current_row += 1
                        continue

                    # Обработка заголовков
                    if line.startswith("===") or line.startswith("ОТЧЕТ"):
                        cell = ws.cell(row=current_row, column=1, value=line.replace("===", "").strip())
                        cell.font = Font(bold=True, size=14)
                        current_row += 2
                        continue

                    # Обработка подзаголовков
                    if ":" in line and not line.startswith("-"):
                        parts = line.split(":", 1)
                        ws.cell(row=current_row, column=1, value=parts[0].strip()).font = Font(bold=True)
                        ws.cell(row=current_row, column=2, value=parts[1].strip())
                        current_row += 1
                        continue

                    # Обработка пунктов списка
                    if line.startswith("-"):
                        ws.cell(row=current_row, column=2, value=line[1:].strip())
                        current_row += 1
                        continue

                    # Обычный текст
                    ws.cell(row=current_row, column=1, value=line)
                    current_row += 1

                # Настройка ширины столбцов
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width

                # Сохраняем файл
                wb.save(file_name)

                QMessageBox.information(
                    self,
                    "Экспорт завершен",
                    f"Отчет успешно экспортирован в {file_name}",
                    QMessageBox.StandardButton.Ok
                )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Ошибка экспорта",
                    f"Не удалось экспортировать отчет: {str(e)}",
                    QMessageBox.StandardButton.Ok
                )

    def generate_report_from_data(self):
        """Генерирует отчет на основе данных из других вкладок"""
        report_type = self.report_type_combo.currentText()
        date_from = self.report_date_from.date()
        date_to = self.report_date_to.date()

        try:
            if report_type == "Отчет по продажам":
                report_text = self.generate_sales_report(date_from, date_to)
            elif report_type == "Отчет по возвратам":
                report_text = self.generate_returns_report(date_from, date_to)
            elif report_type == "Отчет по товарам":
                report_text = self.generate_products_report()
            elif report_type == "Отчет по клиентам":
                report_text = self.generate_clients_report()
            elif report_type == "Сводный отчет":
                report_text = self.generate_summary_report(date_from, date_to)
            else:
                report_text = "Неизвестный тип отчета"

            self.report_preview.setPlainText(report_text)

        except Exception as e:
            self.report_preview.setPlainText(f"Ошибка при формировании отчета: {str(e)}")

    def generate_sales_report(self, date_from, date_to):
        """Генерирует отчет по продажам на основе данных из вкладки заказов"""
        if not hasattr(self, 'all_orders'):
            return "Нет данных о заказах. Сначала загрузите данные на вкладке 'Заказы'."

        # Фильтруем заказы по дате
        filtered_orders = []
        for order in self.all_orders:
            order_date_str = order.get('order_date', '')
            if order_date_str:
                try:
                    order_date = datetime.fromisoformat(order_date_str.replace('Z', '+00:00')).date()
                    qdate = QDate(order_date.year, order_date.month, order_date.day)
                    if date_from <= qdate <= date_to:
                        filtered_orders.append(order)
                except ValueError:
                    continue

        if not filtered_orders:
            return f"Нет данных о продажах за период с {date_from.toString('dd.MM.yyyy')} по {date_to.toString('dd.MM.yyyy')}"

        # Формируем отчет
        report = "ОТЧЕТ О ПРОДАЖАХ\n\n"
        report += f"Период: {date_from.toString('dd.MM.yyyy')} - {date_to.toString('dd.MM.yyyy')}\n"
        report += f"Всего заказов: {len(filtered_orders)}\n"

        # Статистика по статусам
        status_stats = {}
        for order in filtered_orders:
            status = order.get('status', 'Неизвестно')
            status_stats[status] = status_stats.get(status, 0) + 1

        report += "\nСТАТУСЫ ЗАКАЗОВ:\n"
        for status, count in status_stats.items():
            report += f"- {status}: {count} зак. ({count / len(filtered_orders):.1%})\n"

        return report

    def generate_returns_report(self, date_from, date_to):
        """Генерирует отчет по возвратам на основе данных из вкладки возвратов"""
        if not hasattr(self, 'all_returns'):
            return "Нет данных о возвратах. Сначала загрузите данные на вкладке 'Возвраты'."

        # Фильтруем возвраты по дате
        filtered_returns = []
        for return_item in self.all_returns:
            return_date_str = return_item.get('return_date', '')
            if return_date_str:
                try:
                    return_date = datetime.fromisoformat(return_date_str.replace('Z', '+00:00')).date()
                    qdate = QDate(return_date.year, return_date.month, return_date.day)
                    if date_from <= qdate <= date_to:
                        filtered_returns.append(return_item)
                except ValueError:
                    continue

        if not filtered_returns:
            return f"Нет данных о возвратах за период с {date_from.toString('dd.MM.yyyy')} по {date_to.toString('dd.MM.yyyy')}"

        # Формируем отчет
        report = "ОТЧЕТ О ВОЗВРАТАХ\n\n"
        report += f"Период: {date_from.toString('dd.MM.yyyy')} - {date_to.toString('dd.MM.yyyy')}\n"
        report += f"Всего возвратов: {len(filtered_returns)}\n"

        # Статистика по статусам
        status_stats = {}
        reason_stats = {}

        for return_item in filtered_returns:
            status = return_item.get('status', 'Неизвестно')
            reason = return_item.get('description', 'Причина не указана')

            status_stats[status] = status_stats.get(status, 0) + 1
            reason_stats[reason] = reason_stats.get(reason, 0) + 1

        report += "\nСТАТУСЫ ВОЗВРАТОВ:\n"
        for status, count in status_stats.items():
            report += f"- {status}: {count} возв. ({count / len(filtered_returns):.1%})\n"

        report += "\nПРИЧИНЫ ВОЗВРАТОВ:\n"
        for reason, count in reason_stats.items():
            report += f"- {reason}: {count} случаев\n"

        return report

    def generate_products_report(self):
        """Генерирует отчет по товарам на основе данных из вкладки товаров"""
        if not hasattr(self, 'products_table') or self.products_table.rowCount() == 0:
            return "Нет данных о товарах. Сначала загрузите данные на вкладке 'Товары'."

        total_products = self.products_table.rowCount()
        total_stock = 0
        type_stats = {}
        material_stats = {}

        for row in range(self.products_table.rowCount()):
            # Получаем данные из таблицы
            product_type = self.products_table.item(row, 3).text()
            material = self.products_table.item(row, 4).text()
            stock = int(self.products_table.item(row, 8).text())

            total_stock += stock
            type_stats[product_type] = type_stats.get(product_type, 0) + 1
            material_stats[material] = material_stats.get(material, 0) + 1

        # Формируем отчет
        report = "ОТЧЕТ ПО ТОВАРАМ\n\n"
        report += f"Всего товаров: {total_products}\n"
        report += f"Общее количество на складе: {total_stock}\n"

        report += "\nРАСПРЕДЕЛЕНИЕ ПО ТИПАМ:\n"
        for product_type, count in type_stats.items():
            report += f"- {product_type}: {count} шт. ({count / total_products:.1%})\n"

        report += "\nРАСПРЕДЕЛЕНИЕ ПО МАТЕРИАЛАМ:\n"
        for material, count in material_stats.items():
            report += f"- {material}: {count} шт. ({count / total_products:.1%})\n"

        return report

    def generate_clients_report(self):
        """Генерирует отчет по клиентам на основе данных из вкладки пользователей"""
        if not hasattr(self, 'users_table') or self.users_table.rowCount() == 0:
            return "Нет данных о пользователях. Сначала загрузите данные на вкладке 'Пользователи'."

        total_users = self.users_table.rowCount()
        role_stats = {}

        for row in range(self.users_table.rowCount()):
            role = self.users_table.item(row, 2).text()
            role_stats[role] = role_stats.get(role, 0) + 1

        # Формируем отчет
        report = "ОТЧЕТ ПО ПОЛЬЗОВАТЕЛЯМ\n\n"
        report += f"Всего пользователей: {total_users}\n"

        report += "\nРАСПРЕДЕЛЕНИЕ ПО РОЛЯМ:\n"
        for role, count in role_stats.items():
            report += f"- {role}: {count} чел. ({count / total_users:.1%})\n"

        return report

    def generate_summary_report(self, date_from, date_to):
        """Генерирует сводный отчет по всем данным"""
        sales_report = self.generate_sales_report(date_from, date_to)
        returns_report = self.generate_returns_report(date_from, date_to)
        products_report = self.generate_products_report()
        clients_report = self.generate_clients_report()

        report = "СВОДНЫЙ ОТЧЕТ\n\n"
        report += f"Период: {date_from.toString('dd.MM.yyyy')} - {date_to.toString('dd.MM.yyyy')}\n\n"

        report += "=== ПРОДАЖИ ===\n"
        report += sales_report.split('\n', 2)[-1] + "\n\n"

        report += "=== ВОЗВРАТЫ ===\n"
        report += returns_report.split('\n', 2)[-1] + "\n\n"

        report += "=== ТОВАРЫ ===\n"
        report += products_report.split('\n', 2)[-1] + "\n\n"

        report += "=== КЛИЕНТЫ ===\n"
        report += clients_report.split('\n', 2)[-1]

        return report

    def format_returns_report(self, data):
        """Форматирует данные о возвратах в текстовый отчет"""
        report = "ОТЧЕТ О ВОЗВРАТАХ\n\n"
        report += f"Период: {data.get('date_from')} - {data.get('date_to')}\n"
        report += f"Всего возвратов: {data.get('total_returns', 0)}\n"
        report += f"Общая сумма: {data.get('total_amount', 0):,.2f} ₽\n\n"

        report += "СТАТУСЫ ВОЗВРАТОВ:\n"
        for status in data.get('status_stats', []):
            report += f"{status['status']}: {status['count']} ({status['percentage']}%)\n"

        report += "\nПРИЧИНЫ ВОЗВРАТОВ:\n"
        for reason in data.get('reason_stats', []):
            report += f"- {reason['reason']}: {reason['count']} случаев\n"

        return report

    def export_report(self):
        """Экспортирует отчет в выбранный формат"""
        options = QFileDialog.Option(0)
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт отчета",
            "",
            "PDF Files (*.pdf);;Text Files (*.txt);;CSV Files (*.csv)",
            options=options
        )

        if file_name:
            try:
                report_text = self.report_preview.toPlainText()

                if file_name.endswith('.txt'):
                    with open(file_name, 'w', encoding='utf-8') as f:
                        f.write(report_text)
                elif file_name.endswith('.csv'):
                    # Здесь можно добавить логику преобразования в CSV
                    with open(file_name, 'w', encoding='utf-8') as f:
                        f.write(report_text.replace('\n', ','))
                elif file_name.endswith('.pdf'):
                    # Для PDF потребуется дополнительная библиотека, например ReportLab
                    pass

                QMessageBox.information(
                    self,
                    "Экспорт завершен",
                    f"Отчет успешно экспортирован в {file_name}",
                    QMessageBox.StandardButton.Ok
                )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Ошибка экспорта",
                    f"Не удалось экспортировать отчет: {str(e)}",
                    QMessageBox.StandardButton.Ok
                )

    def generate_report(self):
        report_type = self.report_type_combo.currentText()
        date_from = self.report_date_from.date().toString("yyyy-MM-dd")
        date_to = self.report_date_to.date().toString("yyyy-MM-dd")
        self.log_action(f"Формирование отчета: {report_type} за период {date_from} - {date_to}")

        try:
            if report_type == "Отчет по продажам":
                response = requests.get(
                    f"{self.base_url}/reports/sales",
                    params={
                        "date_from": date_from,
                        "date_to": date_to
                    }
                )

                if response.status_code == 200:
                    sales_data = response.json()
                    report_text = self.format_sales_report(sales_data)
                    self.log_action("Отчет по продажам успешно сформирован")
                else:
                    report_text = f"Ошибка при получении данных: {response.text}"
                    self.log_action(f"Ошибка формирования отчета по продажам: {response.text}")

            elif report_type == "Отчет по возвратам":
                response = requests.get(
                    f"{self.base_url}/reports/returns",
                    params={
                        "date_from": date_from,
                        "date_to": date_to
                    }
                )

                if response.status_code == 200:
                    returns_data = response.json()
                    report_text = self.format_returns_report(returns_data)
                    self.log_action("Отчет по возвратам успешно сформирован")
                else:
                    report_text = f"Ошибка при получении данных: {response.text}"
                    self.log_action(f"Ошибка формирования отчета по возвратам: {response.text}")

            else:
                report_text = f"Отчет '{report_type}' за период с {date_from} по {date_to}\n\n"
                report_text += "Функциональность отчета в разработке."
                self.log_action(f"Формирование отчета '{report_type}' (в разработке)")

            self.report_preview.setPlainText(report_text)

        except Exception as e:
            self.log_action(f"Ошибка при формировании отчета: {str(e)}")
            self.report_preview.setPlainText(f"Ошибка при формировании отчета: {str(e)}")

    def format_sales_report(self, data):
        """Форматирует данные о продажах в текстовый отчет"""
        report = "ОТЧЕТ О ПРОДАЖАХ\n\n"
        report += f"Период: {data.get('date_from')} - {data.get('date_to')}\n"
        report += f"Всего заказов: {data.get('total_orders', 0)}\n"
        report += f"Общая сумма: {data.get('total_amount', 0):,.2f} ₽\n\n"

        report += "СТАТИСТИКА ПО ДНЯМ:\n"
        for day in data.get('daily_stats', []):
            report += f"{day['date']}: {day['orders']} зак. на {day['amount']:,.2f} ₽\n"

        report += "\nПОПУЛЯРНЫЕ ТОВАРЫ:\n"
        for i, product in enumerate(data.get('top_products', []), 1):
            report += f"{i}. {product['name']} - {product['sales']} шт. на {product['amount']:,.2f} ₽\n"

        return report

    def update_report_filters(self, report_type):
        """Обновляет доступные фильтры в зависимости от типа отчета"""
        self.specific_filter_combo.clear()

        if report_type == "Отчет по продажам":
            self.specific_filter_combo.setVisible(True)
            self.specific_filter_combo.addItems([
                "Все продажи",
                "По статусу заказа",
                "По типу товара",
                "По менеджеру"
            ])
        elif report_type == "Отчет по возвратам":
            self.specific_filter_combo.setVisible(True)
            self.specific_filter_combo.addItems([
                "Все возвраты",
                "По статусу возврата",
                "По причине возврата"
            ])
        elif report_type == "Отчет по товарам":
            self.specific_filter_combo.setVisible(True)
            self.specific_filter_combo.addItems([
                "Все товары",
                "По наличию на складе",
                "По типу товара",
                "По материалу"
            ])
        else:
            self.specific_filter_combo.setVisible(False)

    def update_details_table(self):
        """Обновляет таблицу с детализированной статистикой"""
        metric = self.metric_combo.currentText()
        self.details_table.setRowCount(0)

        # Заглушка с тестовыми данными
        if metric == "Популярные товары":
            products = [
                {"name": "Золотое кольцо", "sales": 8, "revenue": "64,000 ₽", "share": "32%"},
                {"name": "Серебряные серьги", "sales": 5, "revenue": "25,000 ₽", "share": "20%"},
                {"name": "Платиновый браслет", "sales": 3, "revenue": "45,000 ₽", "share": "18%"},
                {"name": "Золотая цепочка", "sales": 2, "revenue": "14,500 ₽", "share": "12%"},
            ]

            for product in products:
                row = self.details_table.rowCount()
                self.details_table.insertRow(row)

                self.details_table.setItem(row, 0, QTableWidgetItem(product["name"]))
                self.details_table.setItem(row, 1, QTableWidgetItem(str(product["sales"])))
                self.details_table.setItem(row, 2, QTableWidgetItem(product["revenue"]))
                self.details_table.setItem(row, 3, QTableWidgetItem(product["share"]))
        else:
            # Для других метрик - временные данные
            dates = ["01.06.2023", "02.06.2023", "03.06.2023", "04.06.2023", "05.06.2023"]
            values = ["12,000 ₽", "15,500 ₽", "18,200 ₽", "14,800 ₽", "16,300 ₽"]
            changes = ["+5%", "+12%", "+8%", "-4%", "+3%"]
            details = ["5 заказов", "7 заказов", "9 заказов", "6 заказов", "8 заказов"]

            for i in range(len(dates)):
                row = self.details_table.rowCount()
                self.details_table.insertRow(row)

                self.details_table.setItem(row, 0, QTableWidgetItem(dates[i]))
                self.details_table.setItem(row, 1, QTableWidgetItem(values[i]))

                change_item = QTableWidgetItem(changes[i])
                if changes[i].startswith("-"):
                    change_item.setForeground(QColor("#ff5555"))
                else:
                    change_item.setForeground(QColor("#55ff55"))
                self.details_table.setItem(row, 2, change_item)

                self.details_table.setItem(row, 3, QTableWidgetItem(details[i]))

    def create_summary_card(self, title, value, change, icon_path):
        """Создает карточку с краткой статистикой"""
        card = QFrame()
        card.setObjectName("statCard")
        card.setMinimumHeight(100)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(20, 15, 20, 15)
        card_layout.setSpacing(10)

        # Header with icon
        header_layout = QHBoxLayout()

        icon_label = QLabel()
        icon_pixmap = QPixmap(icon_path).scaled(24, 24, Qt.AspectRatioMode.KeepAspectRatio,
                                                Qt.TransformationMode.SmoothTransformation)
        icon_label.setPixmap(icon_pixmap)

        title_label = QLabel(title)
        title_label.setFont(QFont('Montserrat', 14))
        title_label.setStyleSheet("color: #d4af37;")

        header_layout.addWidget(icon_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        card_layout.addLayout(header_layout)

        # Value
        value_label = QLabel(value)
        value_label.setFont(QFont('Montserrat', 24, QFont.Weight.Bold))
        value_label.setStyleSheet("color: white;")
        card_layout.addWidget(value_label)

        # Change indicator
        change_label = QLabel(change)
        change_label.setFont(QFont('Montserrat', 12))

        # Set color based on change value
        if change.startswith("-"):
            change_label.setStyleSheet("color: #ff5555;")
        else:
            change_label.setStyleSheet("color: #55ff55;")

        card_layout.addWidget(change_label)

        return card

    def show_add_user_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить пользователя")
        dialog.setFixedSize(400, 300)

        # Устанавливаем стиль для диалога
        dialog.setStyleSheet("""
            QDialog {
                background-color: #1a1a1a;
            }
            QLabel {
                color: white;
            }
            QLineEdit, QComboBox {
                padding: 10px;
                font-size: 14px;
                border: 1px solid #444;
                border-radius: 4px;
                background: #333;
                color: white;
            }
            #saveButton {
                background-color: #d4af37;
                color: black;
                font-size: 16px;
                padding: 12px;
                border-radius: 5px;
            }
            #saveButton:hover {
                background-color: #f0d87c;
            }
            #cancelButton {
                background-color: #333;
                color: white;
                font-size: 16px;
                padding: 12px;
                border-radius: 5px;
            }
            #cancelButton:hover {
                background-color: #444;
            }
        """)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        title = QLabel("Добавление нового пользователя")
        title.setFont(QFont('Montserrat', 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #d4af37;")
        layout.addWidget(title)

        # Form
        form_layout = QFormLayout()
        form_layout.setHorizontalSpacing(15)
        form_layout.setVerticalSpacing(10)

        # Создаем белые label'ы
        username_label = QLabel("Имя пользователя:")
        password_label = QLabel("Пароль:")
        role_label = QLabel("Роль:")

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Введите имя пользователя")

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Введите пароль")
        self.password_input.setEchoMode(QLineEdit.Password)

        # Создаем комбобокс с ролями (текст для интерфейса)
        self.role_combo = QComboBox()
        self.role_combo.addItems(["Админ", "Ювелир", "Кладовщик", "Клиент"])

        form_layout.addRow(username_label, self.username_input)
        form_layout.addRow(password_label, self.password_input)
        form_layout.addRow(role_label, self.role_combo)

        layout.addLayout(form_layout)

        # Buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setObjectName("cancelButton")
        cancel_btn.clicked.connect(dialog.reject)

        save_btn = QPushButton("Сохранить")
        save_btn.setObjectName("saveButton")
        save_btn.clicked.connect(lambda: self.save_new_user(dialog))

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(save_btn)
        layout.addLayout(buttons_layout)

        # Status
        self.add_user_status = QLabel()
        self.add_user_status.setWordWrap(True)
        self.add_user_status.setStyleSheet("color: white;")
        layout.addWidget(self.add_user_status)

        dialog.exec()

    def save_new_user(self, dialog):
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        role = self.role_combo.currentText()
        self.log_action(f"Создание нового пользователя: {username}")

        if not username or not password:
            self.log_action("Ошибка: не заполнены обязательные поля")
            self.add_user_status.setText("Пожалуйста, заполните все поля")
            self.add_user_status.setStyleSheet("color: #ff5555;")
            return

        try:
            data = {
                "username": username,
                "password": password,
                "role": role
            }

            response = requests.post(
                f"{self.base_url}/create/admin/users",
                json=data,
                headers={"Content-Type": "application/json"}
            )

            if response.status_code == 200:
                self.log_action(f"Пользователь {username} успешно создан")
                self.add_user_status.setText("Пользователь успешно добавлен")
                self.add_user_status.setStyleSheet("color: #55ff55;")
                QTimer.singleShot(1500, dialog.close)
                self.load_users()
            else:
                error_msg = response.json().get('detail', response.text)
                self.log_action(f"Ошибка создания пользователя: {error_msg}")
                self.add_user_status.setText(f"Ошибка: {error_msg}")
                self.add_user_status.setStyleSheet("color: #ff5555;")
        except Exception as e:
            self.log_action(f"Ошибка при создании пользователя: {str(e)}")
            self.add_user_status.setText(f"Ошибка соединения: {str(e)}")
            self.add_user_status.setStyleSheet("color: #ff5555;")

    def show_order_details_admin(self):
        # Получаем выбранный заказ
        selected_row = self.orders_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите заказ для просмотра деталей")
            return

        order_id = int(self.orders_table.item(selected_row, 0).text())

        try:
            # Получаем данные заказа с сервера
            response = requests.get(f"{self.base_url}/orders/get/{order_id}")
            if response.status_code != 200:
                QMessageBox.warning(self, "Ошибка", f"Не удалось получить данные заказа: {response.text}")
                return

            order_data = response.json()

            # Создаем диалоговое окно
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Детали заказа #{order_data['order_id']}")
            dialog.setMinimumSize(700, 500)  # Увеличили минимальный размер для лучшего отображения

            layout = QVBoxLayout(dialog)
            layout.setContentsMargins(20, 20, 20, 20)
            layout.setSpacing(15)

            # Order info
            order_info = QLabel(f"Заказ #{order_data['order_id']}")
            order_info.setFont(QFont('Montserrat', 18, QFont.Weight.Bold))
            order_info.setStyleSheet("color: #d4af37;")
            layout.addWidget(order_info)

            # Client info
            client_info = QLabel(f"Клиент: {order_data['username']}")
            client_info.setFont(QFont('Montserrat', 14))
            layout.addWidget(client_info)

            # Status and date
            status_date_layout = QHBoxLayout()

            status = QLabel(f"Статус: {order_data['status']}")
            status.setFont(QFont('Montserrat', 14))

            # Форматирование даты
            order_date = datetime.fromisoformat(order_data['order_date'].replace('Z', '+00:00'))
            formatted_date = order_date.strftime("%d.%m.%Y %H:%M")

            order_date_label = QLabel(f"Дата: {formatted_date}")
            order_date_label.setFont(QFont('Montserrat', 14))

            status_date_layout.addWidget(status)
            status_date_layout.addStretch()
            status_date_layout.addWidget(order_date_label)
            layout.addLayout(status_date_layout)

            # Items
            items_label = QLabel("Товары:")
            items_label.setFont(QFont('Montserrat', 16, QFont.Weight.Bold))
            items_label.setStyleSheet("color: #d4af37;")
            layout.addWidget(items_label)

            # Items table (теперь с 5 колонками - добавлен артикул)
            items_table = QTableWidget()
            items_table.setColumnCount(5)
            items_table.setHorizontalHeaderLabels(["Товар", "Артикул", "Цена", "Количество", "Сумма"])

            # Настраиваем ширину колонок
            items_table.horizontalHeader().setSectionResizeMode(0,
                                                                QHeaderView.ResizeMode.Stretch)  # Название товара растягивается
            items_table.horizontalHeader().setSectionResizeMode(1,
                                                                QHeaderView.ResizeMode.ResizeToContents)  # Артикул по содержимому
            items_table.horizontalHeader().setSectionResizeMode(2,
                                                                QHeaderView.ResizeMode.ResizeToContents)  # Цена по содержимому
            items_table.horizontalHeader().setSectionResizeMode(3,
                                                                QHeaderView.ResizeMode.ResizeToContents)  # Количество по содержимому
            items_table.horizontalHeader().setSectionResizeMode(4,
                                                                QHeaderView.ResizeMode.ResizeToContents)  # Сумма по содержимому

            items_table.verticalHeader().setVisible(False)
            items_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

            # Заполняем таблицу товарами из заказа
            items_table.setRowCount(len(order_data['items']))
            total_amount = 0

            for row, item in enumerate(order_data['items']):
                items_table.setItem(row, 0, QTableWidgetItem(item['product_name']))  # Название товара
                items_table.setItem(row, 1, QTableWidgetItem(item['article']))  # Артикул
                items_table.setItem(row, 2, QTableWidgetItem(f"{item['price']:,} ₽".replace(",", " ")))  # Цена
                items_table.setItem(row, 3, QTableWidgetItem(str(item['quantity'])))  # Количество

                item_total = item['price'] * item['quantity']
                items_table.setItem(row, 4, QTableWidgetItem(f"{item_total:,} ₽".replace(",", " ")))  # Сумма

                total_amount += item_total

            layout.addWidget(items_table)

            # Total
            total = QLabel(f"Итого: {total_amount:,} ₽".replace(",", " "))
            total.setFont(QFont('Montserrat', 16, QFont.Weight.Bold))
            total.setStyleSheet("color: #d4af37;")
            layout.addWidget(total, alignment=Qt.AlignmentFlag.AlignRight)

            # Status change
            status_layout = QHBoxLayout()
            status_layout.addWidget(QLabel("Изменить статус:"))

            status_combo = QComboBox()
            status_combo.addItems(["В обработке", "Завершен", "Отменен"])
            status_combo.setCurrentText(order_data['status'])

            update_btn = QPushButton("Обновить")
            update_btn.setObjectName("updateButton")
            update_btn.clicked.connect(
                lambda: self.update_order_status(order_data['order_id'], status_combo.currentText(), dialog))

            status_layout.addWidget(status_combo)
            status_layout.addWidget(update_btn)
            status_layout.addStretch()
            layout.addLayout(status_layout)

            dialog.setStyleSheet("""
                QDialog {
                    background-color: #1a1a1a;
                }
                QLabel {
                    color: white;
                }
                QTableWidget {
                    background-color: #252525;
                    border: 1px solid #333;
                    color: white;
                }
                QHeaderView::section {
                    background-color: #333;
                    color: white;
                    padding: 5px;
                    border: none;
                }
                #updateButton {
                    background-color: #d4af37;
                    color: black;
                    font-size: 14px;
                    padding: 8px 16px;
                    border-radius: 5px;
                }
                #updateButton:hover {
                    background-color: #f0d87c;
                }
            """)

            dialog.exec()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить детали заказа: {str(e)}")

    def update_order_status(self, order_id, new_status, dialog):
        try:
            self.log_action(f"Обновление статуса заказа ID: {order_id} на '{new_status}'")
            response = requests.patch(
                f"{self.base_url}/orders/{order_id}/status",
                json={"status": new_status},
                headers={"Content-Type": "application/json"}
            )

            if response.status_code == 200:
                self.log_action(f"Статус заказа ID: {order_id} успешно обновлен")
                QMessageBox.information(self, "Успех", "Статус заказа успешно обновлен")
                dialog.close()
                self.load_orders()
            else:
                self.log_action(f"Ошибка обновления статуса заказа ID: {order_id}: {response.text}")
                QMessageBox.warning(self, "Ошибка", f"Не удалось обновить статус: {response.text}")

        except Exception as e:
            self.log_action(f"Ошибка при обновлении статуса заказа ID: {order_id}: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка соединения: {str(e)}")

    def apply_admin_styles(self):
        self.setStyleSheet("""
            #viewImageButton {
                background-color: #d4af37;
                color: black;
                padding: 5px;
                border-radius: 4px;
                min-width: 80px;
            }
            #viewImageButton:hover {
                background-color: #f0d87c;
            }
            QWidget {
                background-color: #1a1a1a;
                font-family: 'Segoe UI';
            }
            #sidebar {
                background-color: #252525;
                border-right: 1px solid #333;
            }
            #navButton {
                background-color: transparent;
                color: white;
                text-align: left;
                padding: 12px 20px;
                font-size: 16px;
                border-radius: 6px;
            }
            #navButton:hover {
                background-color: #333;
            }
            #logoutButton {
                background-color: #333;
                color: white;
                padding: 12px;
                border-radius: 6px;
            }
            #logoutButton:hover {
                background-color: #444;
            }
            #statCard {;
                border-radius: 10px;
                border: 1px solid #333;
            }
            #activityList, #ordersList {
                background-color: #252525;
                border: 1px solid #333;
                border-radius: 5px;
                color: white;
                font-size: 14px;
            }
            #addButton, #saveButton {
                background-color: #d4af37;
                color: black;
                padding: 10px 20px;
                border-radius: 5px;
            }
            #addButton:hover, #saveButton:hover {
                background-color: #f0d87c;
            }
            #searchButton, #exportButton, #generateButton {
                background-color: #333;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
            }
            #searchButton:hover, #exportButton:hover, #generateButton:hover {
                background-color: #444;
            }
            QTableWidget {
                background-color: #252525;
                border: 1px solid #333;
                color: white;
            }
            QHeaderView::section {
                background-color: #333;
                color: white;
                padding: 5px;
                border: none;
            }
            #editButton, #processButton, #addButton {
                background-color: transparent;
                border-radius: 50%;
                padding: 8px;
            }
            #editButton:hover {
                background-color: #333;
            }
            #deleteButton {
                background-color: transparent;
                border-radius: 50%;
                padding: 8px;
            }
            #deleteButton:hover {
                background-color: #ff5555;
            }
            #processButton:hover {
                background-color: #d4af37;
            }
            #addButton:hover {
                background-color: #55ff55;
            }
            #reportButton {
                background-color: #333;
                color: white;
                padding: 15px;
                border-radius: 5px;
                font-size: 14px;
            }
            #reportButton:hover {
                background-color: #444;
            }
            QTextEdit {
                background-color: #252525;
                border: 1px solid #333;
                color: white;
                font-family: 'Courier New';
                font-size: 14px;
            }
            #detailsButton, #updateButton {
                background-color: #d4af37;
                color: black;
                padding: 10px 20px;
                border-radius: 5px;
            }
            #detailsButton:hover, #updateButton:hover {
                background-color: #f0d87c;
            }
            #addButton, #searchButton {
                color: white;
            }
            
            #addButton {
                background-color: #d4af37;
                color: black;
                padding: 8px 16px;
                border-radius: 5px;
            }
            
            #editButton {
                background-color: #333;
                color: white;
                padding: 8px 16px;
                border-radius: 5px;
            }
            
            #deleteButton {
                background-color: #ff5555;
                color: white;
                padding: 8px 16px;
                border-radius: 5px;
            }
            
            #refreshButton {
                background-color: #333;
                color: white;
                padding: 8px 16px;
                border-radius: 5px;
            }
            QComboBox QAbstractItemView {
                color: white;
                background-color: #333;
                selection-background-color: #d4af37;
                selection-color: black;
                outline: none;
            }
        """)


# Пример использования
if __name__ == "__main__":
    app = QApplication(sys.argv)

    # Для тестирования передаем user_id=1 (администратор)
    admin_window = AdminApp(1)
    admin_window.show()

    sys.exit(app.exec())