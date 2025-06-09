import os
import sys
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
import requests
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFrame, QStackedWidget, QSizePolicy,
    QDialog, QTableWidget, QTableWidgetItem, QHeaderView,
    QListWidget, QListWidgetItem, QTextEdit, QLineEdit,
    QComboBox, QDateEdit, QFormLayout, QMessageBox, QScrollArea, QFileDialog, QGroupBox
)
from PySide6.QtCore import Qt, QSize, QTimer, QDate
from PySide6.QtGui import QFont, QIcon, QPixmap, QImage, QColor


class StorekeeperApp(QMainWindow):
    def __init__(self, user_id):
        super().__init__()
        self.user_id = user_id
        self.setWindowTitle("Гусарочка - Кладовщик")
        self.setMinimumSize(1280, 900)
        self.base_url = "http://localhost:8000"  # Замените на ваш базовый URL API

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Sidebar
        self.create_sidebar(main_layout)

        # Content area
        self.content_stack = QStackedWidget()

        # Products management page
        self.products_page = self.create_products_page()
        self.content_stack.addWidget(self.products_page)

        # Returns page
        self.returns_page = self.create_returns_page()
        self.content_stack.addWidget(self.returns_page)

        # Reports page
        self.reports_page = self.create_reports_page()
        self.content_stack.addWidget(self.reports_page)

        main_layout.addWidget(self.content_stack, stretch=5)

        # Apply styles
        self.apply_styles()

    def create_sidebar(self, layout):
        sidebar = QFrame()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(280)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(20, 40, 20, 40)
        sidebar_layout.setSpacing(20)

        # Logo
        logo = QLabel("ГУСАРОЧКА\nКЛАДОВЩИК")
        logo.setFont(QFont('Montserrat', 18, QFont.Weight.Bold))
        logo.setStyleSheet("color: #d4af37;")
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sidebar_layout.addWidget(logo)

        # Navigation buttons
        self.products_btn = QPushButton("Товары")
        self.products_btn.setIcon(QIcon("icons/catalog.png"))
        self.products_btn.setIconSize(QSize(24, 24))
        self.products_btn.setObjectName("navButton")
        self.products_btn.clicked.connect(lambda: self.content_stack.setCurrentIndex(0))

        self.returns_btn = QPushButton("Возвраты")
        self.returns_btn.setIcon(QIcon("icons/returns.png"))
        self.returns_btn.setIconSize(QSize(24, 24))
        self.returns_btn.setObjectName("navButton")
        self.returns_btn.clicked.connect(lambda: self.content_stack.setCurrentIndex(1))

        self.reports_btn = QPushButton("Отчеты")
        self.reports_btn.setIcon(QIcon("icons/reports.png"))
        self.reports_btn.setIconSize(QSize(24, 24))
        self.reports_btn.setObjectName("navButton")
        self.reports_btn.clicked.connect(lambda: self.content_stack.setCurrentIndex(2))

        sidebar_layout.addWidget(self.products_btn)
        sidebar_layout.addWidget(self.returns_btn)
        sidebar_layout.addWidget(self.reports_btn)
        sidebar_layout.addStretch()

        # Logout button
        logout_btn = QPushButton("Выйти")
        logout_btn.setObjectName("logoutButton")
        logout_btn.clicked.connect(self.close)
        sidebar_layout.addWidget(logout_btn)

        layout.addWidget(sidebar)

    def create_products_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        header = QHBoxLayout()

        title = QLabel("Управление товарами")
        title.setFont(QFont('Montserrat', 20, QFont.Weight.Bold))
        title.setStyleSheet("color: gold;")

        header.addWidget(title)
        header.addStretch()
        layout.addLayout(header)

        # Search bar
        search_frame = QFrame()
        search_frame.setStyleSheet("background-color: #252525; border-radius: 5px;")
        search_layout = QHBoxLayout(search_frame)
        search_layout.setContentsMargins(15, 10, 15, 10)
        search_layout.setSpacing(10)

        self.product_search = QLineEdit()
        self.product_search.setPlaceholderText("Поиск по названию или артикулу...")
        self.product_search.setStyleSheet("""
            QLineEdit {
                background-color: #333;
                color: white;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
            }
        """)

        search_btn = QPushButton("Поиск")
        search_btn.setObjectName("searchButton")
        search_btn.setFixedWidth(100)
        search_btn.clicked.connect(self.search_products)

        search_layout.addWidget(self.product_search)
        search_layout.addWidget(search_btn)
        layout.addWidget(search_frame)

        # Products table
        self.products_table = QTableWidget()
        self.products_table.setColumnCount(12)
        self.products_table.setHorizontalHeaderLabels([
            "ID", "Название", "Артикул", "Тип", "Материал",
            "Вставка", "Вес", "Цена", "На складе",
            "ID изображения", "Дата создания", "Изображение"
        ])
        self.products_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.products_table.verticalHeader().setVisible(False)
        self.products_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.products_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.products_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)

        # Загружаем товары
        self.load_products()
        layout.addWidget(self.products_table, stretch=1)

        # Кнопки управления
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)

        self.add_product_btn = QPushButton("Добавить товар")
        self.add_product_btn.setObjectName("addButton")
        self.add_product_btn.clicked.connect(self.show_add_product_dialog)

        self.edit_product_btn = QPushButton("Редактировать")
        self.edit_product_btn.setObjectName("editButton")
        self.edit_product_btn.clicked.connect(self.edit_selected_product)
        self.edit_product_btn.setEnabled(False)

        refresh_btn = QPushButton("Обновить")
        refresh_btn.setObjectName("refreshButton")
        refresh_btn.clicked.connect(self.load_products)

        buttons_layout.addWidget(self.add_product_btn)
        buttons_layout.addWidget(self.edit_product_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(refresh_btn)
        layout.addLayout(buttons_layout)

        # Подключаем сигнал изменения выделения
        self.products_table.itemSelectionChanged.connect(self.update_product_buttons_state)

        return page

    def show_add_product_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить товар")
        dialog.setFixedSize(600, 800)

        # Основной стиль диалога
        dialog.setStyleSheet("""
            QDialog {
                background-color: #1a1a1a;
                color: white;
                font-family: 'Segoe UI';
            }
            QLabel {
                color: white;
            }
            QLineEdit, QTextEdit {
                background-color: #333;
                color: white;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
                selection-background-color: #d4af37;
                selection-color: black;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 14px;
                min-width: 100px;
            }
            #saveButton {
                background-color: #d4af37;
                color: black;
            }
            #saveButton:hover {
                background-color: #f0d87c;
            }
            #cancelButton {
                background-color: #444;
                color: white;
            }
            #cancelButton:hover {
                background-color: #555;
            }
            #browseButton {
                background-color: #555;
                color: white;
            }
            #browseButton:hover {
                background-color: #666;
            }
            QGroupBox {
                color: #d4af37;
                border: 1px solid #444;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)

        main_layout = QVBoxLayout(dialog)
        main_layout.setContentsMargins(25, 25, 25, 25)
        main_layout.setSpacing(15)

        # Заголовок
        title = QLabel("Добавление нового товара")
        title.setFont(QFont('Montserrat', 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #d4af37; margin-bottom: 15px;")
        main_layout.addWidget(title)

        # Основная форма с полями ввода
        form_scroll = QScrollArea()
        form_scroll.setWidgetResizable(True)
        form_scroll.setStyleSheet("border: none;")

        form_widget = QWidget()
        form_layout = QVBoxLayout(form_widget)
        form_layout.setContentsMargins(5, 5, 5, 5)
        form_layout.setSpacing(15)

        # Поля данных товара
        fields = [
            ("Название товара:", "product_name_input", QLineEdit()),
            ("Артикул:", "product_article_input", QLineEdit()),
            ("Тип:", "product_type_input", QLineEdit()),
            ("Материал:", "product_material_input", QLineEdit()),
            ("Тип вставки:", "product_insert_input", QLineEdit()),
            ("Вес (г):", "product_weight_input", QLineEdit()),
            ("Цена (₽):", "product_price_input", QLineEdit()),
            ("Количество на складе:", "product_quantity_input", QLineEdit())
        ]

        for label_text, attr_name, widget in fields:
            setattr(self, attr_name, widget)
            field_layout = QHBoxLayout()
            field_layout.addWidget(QLabel(label_text), stretch=1)
            field_layout.addWidget(widget, stretch=2)
            form_layout.addLayout(field_layout)

        # Группа для загрузки изображения
        image_group = QGroupBox("Изображение товара")
        image_layout = QVBoxLayout(image_group)
        image_layout.setSpacing(15)

        # Кнопка выбора файла
        browse_btn = QPushButton("Выбрать изображение")
        browse_btn.setObjectName("browseButton")
        browse_btn.clicked.connect(self.select_image_file)
        image_layout.addWidget(browse_btn, alignment=Qt.AlignmentFlag.AlignLeft)

        # Область превью изображения
        self.image_preview = QLabel()
        self.image_preview.setMinimumSize(300, 200)
        self.image_preview.setStyleSheet("""
            QLabel {
                background-color: #333;
                border: 1px dashed #555;
                qproperty-alignment: 'AlignCenter';
            }
        """)
        image_layout.addWidget(self.image_preview)

        # Статус выбранного файла
        self.image_path_label = QLabel("Файл не выбран")
        self.image_path_label.setStyleSheet("color: #888; font-size: 12px;")
        image_layout.addWidget(self.image_path_label, alignment=Qt.AlignmentFlag.AlignCenter)

        form_layout.addWidget(image_group)
        form_scroll.setWidget(form_widget)
        main_layout.addWidget(form_scroll)

        # Кнопки управления внизу
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setObjectName("cancelButton")
        cancel_btn.clicked.connect(dialog.reject)

        save_btn = QPushButton("Сохранить")
        save_btn.setObjectName("saveButton")
        save_btn.clicked.connect(lambda: self.save_new_product_with_image(dialog))

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(save_btn)
        main_layout.addLayout(buttons_layout)

        # Статусная строка
        self.add_product_status = QLabel()
        self.add_product_status.setWordWrap(True)
        self.add_product_status.setStyleSheet("""
            QLabel {
                color: #ff5555;
                font-size: 13px;
                margin-top: 10px;
                padding: 5px;
            }
        """)
        main_layout.addWidget(self.add_product_status)

        self.current_image_path = None
        dialog.exec()

    def select_image_file(self, update_preview=False):
        file_dialog = QFileDialog()
        file_dialog.setStyleSheet("""
            QFileDialog {
                background-color: #1a1a1a;
                color: white;
            }
            QLabel {
                color: white;
            }
            QPushButton {
                background-color: #444;
                color: white;
                padding: 5px 10px;
                border-radius: 4px;
            }
        """)
        file_dialog.setNameFilter("Images (*.png *.jpg *.jpeg *.bmp)")
        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            if selected_files:
                self.current_image_path = selected_files[0]
                self.image_path_label.setText(os.path.basename(self.current_image_path))
                self.image_path_label.setStyleSheet("color: #d4af37; font-size: 12px;")

                pixmap = QPixmap(self.current_image_path)
                if not pixmap.isNull():
                    pixmap = pixmap.scaled(
                        300, 200,
                        Qt.AspectRatioMode.KeepAspectRatio,
                        Qt.TransformationMode.SmoothTransformation
                    )
                    self.image_preview.setPixmap(pixmap)

    def save_new_product_with_image(self, dialog):
        if not all([
            self.product_name_input.text().strip(),
            self.product_article_input.text().strip(),
            self.product_type_input.text().strip(),
            self.product_material_input.text().strip(),
            self.product_weight_input.text().strip(),
            self.product_price_input.text().strip(),
            self.product_quantity_input.text().strip()
        ]):
            self.add_product_status.setText("Заполните все обязательные поля")
            self.add_product_status.setStyleSheet("color: #ff5555;")
            return

        try:
            image_id = 0
            if self.current_image_path:
                with open(self.current_image_path, 'rb') as img_file:
                    files = {'file': (self.current_image_path.split('/')[-1], img_file, 'image/jpeg')}
                    response = requests.post(
                        f"{self.base_url}/images/upload/",
                        files=files
                    )
                    if response.status_code == 200:
                        image_id = response.json().get('id', 0)
                    else:
                        raise Exception(f"Ошибка загрузки изображения: {response.text}")

            product_data = {
                "name": self.product_name_input.text().strip(),
                "article": self.product_article_input.text().strip(),
                "type": self.product_type_input.text().strip(),
                "material": self.product_material_input.text().strip(),
                "insert_type": self.product_insert_input.text().strip(),
                "weight": float(self.product_weight_input.text()),
                "price": float(self.product_price_input.text()),
                "stock_quantity": int(self.product_quantity_input.text()),
                "image_id": image_id
            }

            response = requests.post(
                f"{self.base_url}/products/create",
                json=product_data,
                headers={"Content-Type": "application/json"}
            )

            if response.status_code == 200:
                self.add_product_status.setText("Товар успешно добавлен")
                self.add_product_status.setStyleSheet("color: #55ff55;")
                QTimer.singleShot(1500, dialog.close)
                self.load_products()
            else:
                self.add_product_status.setText(f"Ошибка при добавлении товара: {response.text}")
                self.add_product_status.setStyleSheet("color: #ff5555;")

        except Exception as e:
            self.add_product_status.setText(f"Ошибка: {str(e)}")
            self.add_product_status.setStyleSheet("color: #ff5555;")

    def search_products(self):
        search_text = self.product_search.text().strip().lower()

        try:
            response = requests.get(f"{self.base_url}/products/get/all/")
            if response.status_code == 200:
                products = response.json()
                filtered_products = []

                for product in products:
                    matches = [
                        search_text in product.get('name', '').lower(),
                        search_text in product.get('article', '').lower()
                    ]

                    if any(matches):
                        filtered_products.append(product)

                self.populate_products_table(filtered_products)
            else:
                QMessageBox.warning(self, "Ошибка", f"Не удалось выполнить поиск. Код ошибки: {response.status_code}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при поиске товаров: {str(e)}")

    def update_product_buttons_state(self):
        selected = self.products_table.selectionModel().hasSelection()
        self.edit_product_btn.setEnabled(selected)

    def edit_selected_product(self):
        selected_row = self.products_table.currentRow()
        if selected_row >= 0:
            product_id = int(self.products_table.item(selected_row, 0).text())
            product_data = {
                "product_id": product_id,
                "name": self.products_table.item(selected_row, 1).text(),
                "article": self.products_table.item(selected_row, 2).text(),
                "type": self.products_table.item(selected_row, 3).text(),
                "material": self.products_table.item(selected_row, 4).text(),
                "insert_type": self.products_table.item(selected_row, 5).text(),
                "weight": float(self.products_table.item(selected_row, 6).text()),
                "price": float(self.products_table.item(selected_row, 7).text().replace(" ₽", "").replace(" ", "")),
                "stock_quantity": int(self.products_table.item(selected_row, 8).text()),
                "image_id": int(self.products_table.item(selected_row, 9).text()) if self.products_table.item(
                    selected_row, 9) else 0
            }
            self.show_edit_product_dialog(product_data)

    def show_edit_product_dialog(self, product_data):
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Редактирование товара ID: {product_data['product_id']}")
        dialog.setFixedSize(600, 800)

        dialog.setStyleSheet("""
            QDialog {
                background-color: #1a1a1a;
                color: white;
                font-family: 'Segoe UI';
            }
            QLabel {
                color: white;
            }
            QLineEdit, QTextEdit {
                background-color: #333;
                color: white;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
                selection-background-color: #d4af37;
                selection-color: black;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 14px;
                min-width: 100px;
            }
            #saveButton {
                background-color: #d4af37;
                color: black;
            }
            #saveButton:hover {
                background-color: #f0d87c;
            }
            #cancelButton {
                background-color: #444;
                color: white;
            }
            #cancelButton:hover {
                background-color: #555;
            }
            #browseButton {
                background-color: #555;
                color: white;
            }
            #browseButton:hover {
                background-color: #666;
            }
            QGroupBox {
                color: #d4af37;
                border: 1px solid #444;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)

        self.original_product_data = product_data.copy()
        self.current_image_path = None
        self.new_image_id = None

        main_layout = QVBoxLayout(dialog)
        main_layout.setContentsMargins(25, 25, 25, 25)
        main_layout.setSpacing(15)

        title = QLabel(f"Редактирование товара: {product_data['name']}")
        title.setFont(QFont('Montserrat', 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #d4af37; margin-bottom: 15px;")
        main_layout.addWidget(title)

        form_scroll = QScrollArea()
        form_scroll.setWidgetResizable(True)
        form_scroll.setStyleSheet("border: none;")

        form_widget = QWidget()
        form_layout = QVBoxLayout(form_widget)
        form_layout.setContentsMargins(5, 5, 5, 5)
        form_layout.setSpacing(15)

        fields = [
            ("Название:", "edit_name_input", QLineEdit(product_data["name"])),
            ("Артикул:", "edit_article_input", QLineEdit(product_data["article"])),
            ("Тип:", "edit_type_input", QLineEdit(product_data["type"])),
            ("Материал:", "edit_material_input", QLineEdit(product_data["material"])),
            ("Тип вставки:", "edit_insert_input", QLineEdit(product_data["insert_type"])),
            ("Вес (г):", "edit_weight_input", QLineEdit(str(product_data["weight"]))),
            ("Цена (₽):", "edit_price_input", QLineEdit(str(product_data["price"]))),
            ("Количество:", "edit_quantity_input", QLineEdit(str(product_data["stock_quantity"])))
        ]

        for label_text, attr_name, widget in fields:
            setattr(self, attr_name, widget)
            field_layout = QHBoxLayout()
            field_layout.addWidget(QLabel(label_text), stretch=1)
            field_layout.addWidget(widget, stretch=2)
            form_layout.addLayout(field_layout)

        image_group = QGroupBox("Изображение товара")
        image_layout = QVBoxLayout(image_group)
        image_layout.setSpacing(15)

        self.edit_image_path_label = QLabel("Файл не выбран")
        self.edit_image_path_label.setStyleSheet("color: #888; font-size: 12px;")

        self.edit_image_preview = QLabel()
        self.edit_image_preview.setMinimumSize(300, 200)
        self.edit_image_preview.setStyleSheet("""
            QLabel {
                background-color: #333;
                border: 1px dashed #555;
                qproperty-alignment: 'AlignCenter';
            }
        """)

        if product_data["image_id"]:
            self.load_product_image_preview(product_data["image_id"], self.edit_image_preview)
            self.edit_image_status = QLabel(f"Текущее изображение (ID: {product_data['image_id']})")
        else:
            self.edit_image_status = QLabel("Изображение отсутствует")

        self.edit_image_status.setStyleSheet("color: #888; font-size: 12px;")

        browse_btn = QPushButton("Изменить изображение")
        browse_btn.setObjectName("browseButton")
        browse_btn.clicked.connect(self.select_edit_image_file)

        image_layout.addWidget(browse_btn, alignment=Qt.AlignmentFlag.AlignLeft)
        image_layout.addWidget(self.edit_image_preview)
        image_layout.addWidget(self.edit_image_path_label, alignment=Qt.AlignmentFlag.AlignCenter)
        image_layout.addWidget(self.edit_image_status, alignment=Qt.AlignmentFlag.AlignCenter)

        form_layout.addWidget(image_group)
        form_scroll.setWidget(form_widget)
        main_layout.addWidget(form_scroll)

        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setObjectName("cancelButton")
        cancel_btn.clicked.connect(dialog.reject)

        save_btn = QPushButton("Сохранить изменения")
        save_btn.setObjectName("saveButton")
        save_btn.clicked.connect(lambda: self.update_product(product_data["product_id"], dialog))

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(save_btn)
        main_layout.addLayout(buttons_layout)

        self.edit_status_label = QLabel()
        self.edit_status_label.setWordWrap(True)
        self.edit_status_label.setStyleSheet("""
            QLabel {
                color: #ff5555;
                font-size: 13px;
                margin-top: 10px;
                padding: 5px;
            }
        """)
        main_layout.addWidget(self.edit_status_label)

        dialog.exec()

    def select_edit_image_file(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Images (*.png *.jpg *.jpeg *.bmp)")
        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            if selected_files:
                self.current_image_path = selected_files[0]
                self.edit_image_path_label.setText(os.path.basename(self.current_image_path))
                self.edit_image_path_label.setStyleSheet("color: #d4af37;")

                pixmap = QPixmap(self.current_image_path)
                if not pixmap.isNull():
                    pixmap = pixmap.scaled(300, 200, Qt.AspectRatioMode.KeepAspectRatio)
                    self.edit_image_preview.setPixmap(pixmap)
                    self.edit_image_status.setText("Новое изображение (не сохранено)")

    def update_product(self, product_id, dialog):
        try:
            update_data = {
                "name": self.edit_name_input.text().strip(),
                "article": self.edit_article_input.text().strip(),
                "type": self.edit_type_input.text().strip(),
                "material": self.edit_material_input.text().strip(),
                "insert_type": self.edit_insert_input.text().strip(),
                "weight": float(self.edit_weight_input.text()),
                "price": float(self.edit_price_input.text()),
                "stock_quantity": int(self.edit_quantity_input.text()),
                "image_id": self.original_product_data["image_id"]
            }

            if self.current_image_path:
                with open(self.current_image_path, 'rb') as img_file:
                    files = {'file': (os.path.basename(self.current_image_path), img_file, 'image/jpeg')}
                    upload_response = requests.post(
                        f"{self.base_url}/images/upload/",
                        files=files
                    )

                    if upload_response.status_code != 200:
                        raise Exception(f"Ошибка загрузки изображения: {upload_response.text}")

                    new_image_id = upload_response.json().get('id')
                    if not new_image_id:
                        raise Exception("Не получен ID нового изображения")

                    update_data["image_id"] = new_image_id

            update_response = requests.patch(
                f"{self.base_url}/products/update/{product_id}",
                json=update_data,
                headers={"Content-Type": "application/json"}
            )

            if update_response.status_code == 200:
                self.edit_status_label.setText("Товар успешно обновлен")
                self.edit_status_label.setStyleSheet("color: #55ff55;")
                QTimer.singleShot(1500, lambda: (dialog.close(), self.load_products()))
            else:
                error_msg = update_response.json().get('detail', update_response.text)
                raise Exception(f"Ошибка обновления товара: {error_msg}")

        except Exception as e:
            self.edit_status_label.setText(f"Ошибка: {str(e)}")
            self.edit_status_label.setStyleSheet("color: #ff5555;")
            if hasattr(self, 'original_product_data') and self.original_product_data.get("image_id"):
                self.load_product_image_preview(self.original_product_data["image_id"], self.edit_image_preview)
                self.edit_image_status.setText(f"Текущее изображение (ID: {self.original_product_data['image_id']})")

    def load_product_image_preview(self, image_id, preview_label):
        try:
            response = requests.get(f"{self.base_url}/images/{image_id}")
            if response.status_code == 200:
                pixmap = QPixmap()
                pixmap.loadFromData(response.content)
                if not pixmap.isNull():
                    pixmap = pixmap.scaled(300, 200, Qt.AspectRatioMode.KeepAspectRatio)
                    preview_label.setPixmap(pixmap)
        except Exception as e:
            print(f"Ошибка загрузки изображения: {str(e)}")

    def load_products(self):
        try:
            response = requests.get(f"{self.base_url}/products/get/all/")

            if response.status_code == 200:
                products = response.json()
                self.populate_products_table(products)
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить товары. Код ошибки: {response.status_code}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке товаров: {str(e)}")

    def populate_products_table(self, products):
        self.products_table.setRowCount(0)

        for product in products:
            row_position = self.products_table.rowCount()
            self.products_table.insertRow(row_position)

            self.products_table.setItem(row_position, 0, QTableWidgetItem(str(product.get('product_id', ''))))
            self.products_table.setItem(row_position, 1, QTableWidgetItem(product.get('name', '')))
            self.products_table.setItem(row_position, 2, QTableWidgetItem(product.get('article', '')))
            self.products_table.setItem(row_position, 3, QTableWidgetItem(product.get('type', '')))
            self.products_table.setItem(row_position, 4, QTableWidgetItem(product.get('material', '')))
            self.products_table.setItem(row_position, 5, QTableWidgetItem(product.get('insert_type', '')))
            self.products_table.setItem(row_position, 6, QTableWidgetItem(str(product.get('weight', 0))))
            self.products_table.setItem(row_position, 7,
                                        QTableWidgetItem(f"{product.get('price', 0):,} ₽".replace(",", " ")))
            self.products_table.setItem(row_position, 8, QTableWidgetItem(str(product.get('stock_quantity', 0))))
            self.products_table.setItem(row_position, 9, QTableWidgetItem(str(product.get('image_id', 0))))

            created_at = product.get('created_at')
            date_item = QTableWidgetItem()

            if created_at:
                try:
                    if '.' in created_at:
                        dt = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%S.%f")
                    else:
                        dt = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%S")

                    formatted_date = dt.strftime("%d.%m.%Y %H:%M")
                    date_item.setText(formatted_date)
                except Exception as e:
                    print(f"Ошибка форматирования даты {created_at}: {str(e)}")
                    date_item.setText(created_at)
            else:
                date_item.setText("Нет данных")

            self.products_table.setItem(row_position, 10, date_item)

            image_id = product.get('image_id', 0)
            if image_id:
                btn = QPushButton("Просмотр")
                btn.setObjectName("viewImageButton")
                btn.setStyleSheet("""
                            QPushButton {
                                background-color: #d4af37;
                                color: black;
                                padding: 5px;
                                border-radius: 4px;
                            }
                            QPushButton:hover {
                                background-color: #f0d87c;
                            }
                        """)
                btn.clicked.connect(lambda _, img_id=image_id: self.view_product_image(img_id))
                self.products_table.setCellWidget(row_position, 11, btn)
            else:
                self.products_table.setItem(row_position, 11, QTableWidgetItem("Нет изображения"))

    def view_product_image(self, image_id):
        try:
            response = requests.get(f"{self.base_url}/images/{image_id}")

            if response.status_code == 200:
                dialog = QDialog(self)
                dialog.setWindowTitle(f"Изображение товара (ID: {image_id})")
                dialog.setMinimumSize(400, 400)

                layout = QVBoxLayout(dialog)

                image_label = QLabel()
                image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

                image_data = response.content
                pixmap = QPixmap()
                pixmap.loadFromData(image_data)

                if pixmap.width() > 800 or pixmap.height() > 600:
                    pixmap = pixmap.scaled(800, 600, Qt.AspectRatioMode.KeepAspectRatio)

                image_label.setPixmap(pixmap)

                close_btn = QPushButton("Закрыть")
                close_btn.clicked.connect(dialog.close)

                layout.addWidget(image_label)
                layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)

                dialog.exec()
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить изображение. Код ошибки: {response.status_code}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке изображения: {str(e)}")

    def create_returns_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        title = QLabel("Управление возвратами")
        title.setFont(QFont('Montserrat', 20, QFont.Weight.Bold))
        title.setStyleSheet("color: gold;")
        layout.addWidget(title)

        # Filters
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.return_status_combo = QComboBox()
        self.return_status_combo.addItems(["Все статусы", "В обработке", "Одобрен", "Отклонен", "Завершен"])

        filter_layout.addWidget(self.return_status_combo)
        filter_layout.addStretch()

        apply_filters_btn = QPushButton("Применить фильтры")
        apply_filters_btn.setObjectName("applyButton")
        apply_filters_btn.clicked.connect(self.apply_return_filters)
        filter_layout.addWidget(apply_filters_btn)

        layout.addLayout(filter_layout)

        # Returns table
        self.returns_table = QTableWidget()
        self.returns_table.setColumnCount(7)
        self.returns_table.setHorizontalHeaderLabels(
            ["ID возврата", "ID заказа", "Клиент", "Дата", "Статус", "Адрес", "Причина"])
        self.returns_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.returns_table.verticalHeader().setVisible(False)

        # Загружаем возвраты при открытии страницы
        self.load_returns()

        layout.addWidget(self.returns_table)

        # Кнопка изменения статуса
        self.update_status_btn = QPushButton("Изменить статус")
        self.update_status_btn.setObjectName("updateButton")
        self.update_status_btn.clicked.connect(self.update_return_status)
        self.update_status_btn.setEnabled(False)
        layout.addWidget(self.update_status_btn, alignment=Qt.AlignmentFlag.AlignRight)

        # Подключаем сигнал изменения выделения
        self.returns_table.itemSelectionChanged.connect(self.update_return_buttons_state)

        return page

    def apply_return_filters(self):
        status_filter = self.return_status_combo.currentText()

        filtered_returns = []
        for return_item in getattr(self, 'all_returns', []):
            status_match = (status_filter == "Все статусы") or (return_item.get('status', '') == status_filter)

            if status_match:
                filtered_returns.append(return_item)

        self.populate_returns_table(filtered_returns)

    def load_returns(self):
        try:
            response = requests.get(f"{self.base_url}/returns/get/all/")

            if response.status_code == 200:
                returns = response.json()
                self.all_returns = returns
                self.populate_returns_table(returns)
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить возвраты. Код ошибки: {response.status_code}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке возвратов: {str(e)}")

    def populate_returns_table(self, returns):
        self.returns_table.setRowCount(0)

        for return_item in returns:
            row_position = self.returns_table.rowCount()
            self.returns_table.insertRow(row_position)

            self.returns_table.setItem(row_position, 0, QTableWidgetItem(str(return_item.get('return_id', ''))))
            self.returns_table.setItem(row_position, 1, QTableWidgetItem(str(return_item.get('order_id', ''))))
            self.returns_table.setItem(row_position, 2, QTableWidgetItem(str(return_item.get('client_id', ''))))

            return_date = return_item.get('return_date', '')
            if return_date:
                try:
                    dt = datetime.fromisoformat(return_date.replace('Z', '+00:00'))
                    date_str = dt.strftime("%d.%m.%Y %H:%M")
                except ValueError:
                    date_str = return_date
            else:
                date_str = "Нет данных"
            self.returns_table.setItem(row_position, 3, QTableWidgetItem(date_str))

            self.returns_table.setItem(row_position, 4, QTableWidgetItem(return_item.get('status', '')))
            self.returns_table.setItem(row_position, 5, QTableWidgetItem("Адрес не указан"))
            self.returns_table.setItem(row_position, 6, QTableWidgetItem(return_item.get('description', '')))

    def update_return_buttons_state(self):
        selected = self.returns_table.selectionModel().hasSelection()
        self.update_status_btn.setEnabled(selected)

    def update_return_status(self):
        selected_row = self.returns_table.currentRow()
        if selected_row >= 0:
            return_id = int(self.returns_table.item(selected_row, 0).text())
            current_status = self.returns_table.item(selected_row, 4).text()

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Изменение статуса возврата #{return_id}")
            dialog.setFixedSize(400, 200)

            layout = QVBoxLayout(dialog)
            layout.setContentsMargins(20, 20, 20, 20)
            layout.setSpacing(15)

            title = QLabel(f"Изменение статуса возврата #{return_id}")
            title.setFont(QFont('Montserrat', 14, QFont.Weight.Bold))
            title.setStyleSheet("color: #d4af37;")
            layout.addWidget(title)

            status_combo = QComboBox()
            status_combo.addItems(["В обработке", "Одобрен", "Отклонен", "Завершен"])
            status_combo.setCurrentText(current_status)

            form_layout = QFormLayout()
            form_layout.addRow("Текущий статус:", QLabel(current_status))
            form_layout.addRow("Новый статус:", status_combo)
            layout.addLayout(form_layout)

            buttons_layout = QHBoxLayout()
            buttons_layout.addStretch()

            cancel_btn = QPushButton("Отмена")
            cancel_btn.setObjectName("cancelButton")
            cancel_btn.clicked.connect(dialog.reject)

            save_btn = QPushButton("Сохранить")
            save_btn.setObjectName("saveButton")
            save_btn.clicked.connect(lambda: self.save_return_status(return_id, status_combo.currentText(), dialog))

            buttons_layout.addWidget(cancel_btn)
            buttons_layout.addWidget(save_btn)
            layout.addLayout(buttons_layout)

            dialog.setStyleSheet("""
                QDialog {
                    background-color: #1a1a1a;
                }
                QLabel {
                    color: white;
                }
                QComboBox {
                    background-color: #333;
                    color: white;
                    border: 1px solid #444;
                    border-radius: 4px;
                    padding: 8px;
                }
            """)

            dialog.exec()

    def save_return_status(self, return_id, new_status, dialog):
        try:
            response = requests.patch(
                f"{self.base_url}/returns/update-status/{return_id}",
                json={"status": new_status},
                headers={"Content-Type": "application/json"}
            )

            if response.status_code == 200:
                QMessageBox.information(self, "Успех", "Статус возврата успешно обновлен")
                dialog.close()
                self.load_returns()
            else:
                QMessageBox.warning(self, "Ошибка", f"Не удалось обновить статус: {response.text}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка соединения: {str(e)}")

    def create_reports_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        title = QLabel("Отчеты по складу")
        title.setFont(QFont('Montserrat', 20, QFont.Weight.Bold))
        title.setStyleSheet("color: gold;")
        layout.addWidget(title)

        # Report type selection
        report_group = QGroupBox("Тип отчета")
        report_group.setStyleSheet("""
            QGroupBox {
                color: #d4af37;
                border: 1px solid #444;
                border-radius: 5px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)

        report_type_layout = QHBoxLayout(report_group)

        self.report_type_combo = QComboBox()
        self.report_type_combo.addItems([
            "Отчет по остаткам",
            "Отчет по движению товаров"
        ])
        report_type_layout.addWidget(self.report_type_combo)
        report_type_layout.addStretch()

        layout.addWidget(report_group)

        # Buttons
        button_layout = QHBoxLayout()

        generate_btn = QPushButton("Сформировать отчет")
        generate_btn.setObjectName("generateButton")
        generate_btn.clicked.connect(self.generate_report)

        export_btn = QPushButton("Экспорт в Excel")
        export_btn.setObjectName("exportButton")
        export_btn.clicked.connect(self.export_report_to_xlsx)

        button_layout.addWidget(generate_btn)
        button_layout.addWidget(export_btn)
        layout.addLayout(button_layout)

        # Report preview area
        self.report_preview = QTextEdit()
        self.report_preview.setReadOnly(True)
        self.report_preview.setStyleSheet("""
            QTextEdit {
                background-color: #252525;
                border: 1px solid #444;
                border-radius: 5px;
                color: white;
                font-family: 'Consolas';
                font-size: 14px;
            }
        """)
        self.report_preview.setPlainText(
            "Здесь будет отображаться сформированный отчет.\n"
            "Выберите тип отчета и нажмите 'Сформировать отчет'."
        )

        layout.addWidget(self.report_preview, stretch=1)

        return page

    def generate_report(self):
        report_type = self.report_type_combo.currentText()

        try:
            if report_type == "Отчет по остаткам":
                response = requests.get(f"{self.base_url}/products/get/all/")
                if response.status_code == 200:
                    products = response.json()
                    report_text = self.generate_stock_report(products)
                else:
                    report_text = f"Ошибка при получении данных: {response.text}"
            elif report_type == "Отчет по движению товаров":
                report_text = "Отчет по движению товаров (функционал в разработке)"
            else:
                report_text = f"Отчет '{report_type}' не поддерживается"

            self.report_preview.setPlainText(report_text)

        except Exception as e:
            self.report_preview.setPlainText(f"Ошибка при формировании отчета: {str(e)}")

    def generate_stock_report(self, products):
        report = "ОТЧЕТ ПО ОСТАТКАМ НА СКЛАДЕ\n\n"
        report += f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
        report += f"Всего товаров: {len(products)}\n"

        # Общее количество на складе
        total_quantity = sum(product.get('stock_quantity', 0) for product in products)
        report += f"Общее количество на складе: {total_quantity}\n\n"

        # Таблица товаров
        report += "СПИСОК ТОВАРОВ:\n"
        report += "{:<8} {:<30} {:<15} {:<10} {:<10}\n".format(
            "ID", "Название", "Артикул", "Цена", "Кол-во"
        )
        report += "-" * 80 + "\n"

        for product in sorted(products, key=lambda x: x.get('stock_quantity', 0), reverse=True):
            report += "{:<8} {:<30} {:<15} {:<10} {:<10}\n".format(
                str(product.get('product_id', '')),
                product.get('name', '')[:28],
                product.get('article', ''),
                str(product.get('price', 0)),
                str(product.get('stock_quantity', 0))
            )

        # Товары с низким остатком
        low_stock = [p for p in products if p.get('stock_quantity', 0) < 5]
        if low_stock:
            report += "\nТОВАРЫ С НИЗКИМ ОСТАТКОМ (<5):\n"
            for product in low_stock:
                report += f"- {product.get('name', '')} (ID: {product.get('product_id', '')}): " \
                          f"{product.get('stock_quantity', 0)} шт.\n"

        return report

    def export_report_to_xlsx(self):
        report_text = self.report_preview.toPlainText()
        if not report_text or "Здесь будет отображаться" in report_text:
            QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта. Сначала сформируйте отчет.")
            return

        options = QFileDialog.Option(0)
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт отчета в Excel",
            "",
            "Excel Files (*.xlsx)",
            options=options
        )

        if file_name:
            try:
                if not file_name.endswith('.xlsx'):
                    file_name += '.xlsx'

                wb = Workbook()
                ws = wb.active
                ws.title = "Отчет"

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                lines = report_text.split('\n')
                current_row = 1

                for line in lines:
                    if not line.strip():
                        current_row += 1
                        continue

                    if line.startswith("===") or line.startswith("ОТЧЕТ"):
                        cell = ws.cell(row=current_row, column=1, value=line.replace("===", "").strip())
                        cell.font = Font(bold=True, size=14)
                        current_row += 2
                        continue

                    if ":" in line and not line.startswith("-"):
                        parts = line.split(":", 1)
                        ws.cell(row=current_row, column=1, value=parts[0].strip()).font = Font(bold=True)
                        ws.cell(row=current_row, column=2, value=parts[1].strip())
                        current_row += 1
                        continue

                    if line.startswith("-"):
                        ws.cell(row=current_row, column=2, value=line[1:].strip())
                        current_row += 1
                        continue

                    ws.cell(row=current_row, column=1, value=line)
                    current_row += 1

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width

                wb.save(file_name)

                QMessageBox.information(
                    self,
                    "Экспорт завершен",
                    f"Отчет успешно экспортирован в {file_name}",
                    QMessageBox.StandardButton.Ok
                )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Ошибка экспорта",
                    f"Не удалось экспортировать отчет: {str(e)}",
                    QMessageBox.StandardButton.Ok
                )

    def apply_styles(self):
        self.setStyleSheet("""
            #viewImageButton {
                background-color: #d4af37;
                color: black;
                padding: 5px;
                border-radius: 4px;
                min-width: 80px;
            }
            #viewImageButton:hover {
                background-color: #f0d87c;
            }
            QWidget {
                background-color: #1a1a1a;
                font-family: 'Segoe UI';
            }
            #sidebar {
                background-color: #252525;
                border-right: 1px solid #333;
            }
            #navButton {
                background-color: transparent;
                color: white;
                text-align: left;
                padding: 12px 20px;
                font-size: 16px;
                border-radius: 6px;
            }
            #navButton:hover {
                background-color: #333;
            }
            #logoutButton {
                background-color: #333;
                color: white;
                padding: 12px;
                border-radius: 6px;
            }
            #logoutButton:hover {
                background-color: #444;
            }
            #statCard {;
                border-radius: 10px;
                border: 1px solid #333;
            }
            #activityList, #ordersList {
                background-color: #252525;
                border: 1px solid #333;
                border-radius: 5px;
                color: white;
                font-size: 14px;
            }
            #addButton, #saveButton {
                background-color: #d4af37;
                color: black;
                padding: 10px 20px;
                border-radius: 5px;
            }
            #addButton:hover, #saveButton:hover {
                background-color: #f0d87c;
            }
            #searchButton, #exportButton, #generateButton {
                background-color: #333;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
            }
            #searchButton:hover, #exportButton:hover, #generateButton:hover {
                background-color: #444;
            }
            QTableWidget {
                background-color: #252525;
                border: 1px solid #333;
                color: white;
            }
            QHeaderView::section {
                background-color: #333;
                color: white;
                padding: 5px;
                border: none;
            }
            #editButton, #processButton, #addButton {
                background-color: transparent;
                border-radius: 50%;
                padding: 8px;
            }
            #editButton:hover {
                background-color: #333;
            }
            #deleteButton {
                background-color: transparent;
                border-radius: 50%;
                padding: 8px;
            }
            #deleteButton:hover {
                background-color: #ff5555;
            }
            #processButton:hover {
                background-color: #d4af37;
            }
            #addButton:hover {
                background-color: #55ff55;
            }
            #reportButton {
                background-color: #333;
                color: white;
                padding: 15px;
                border-radius: 5px;
                font-size: 14px;
            }
            #reportButton:hover {
                background-color: #444;
            }
            QTextEdit {
                background-color: #252525;
                border: 1px solid #333;
                color: white;
                font-family: 'Courier New';
                font-size: 14px;
            }
            #detailsButton, #updateButton {
                background-color: #d4af37;
                color: black;
                padding: 10px 20px;
                border-radius: 5px;
            }
            #detailsButton:hover, #updateButton:hover {
                background-color: #f0d87c;
            }
            #addButton, #searchButton {
                color: white;
            }

            #addButton {
                background-color: #d4af37;
                color: black;
                padding: 8px 16px;
                border-radius: 5px;
            }

            #editButton {
                background-color: #333;
                color: white;
                padding: 8px 16px;
                border-radius: 5px;
            }

            #deleteButton {
                background-color: #ff5555;
                color: white;
                padding: 8px 16px;
                border-radius: 5px;
            }

            #refreshButton {
                background-color: #333;
                color: white;
                padding: 8px 16px;
                border-radius: 5px;
            }
            QComboBox QAbstractItemView {
                color: white;
                background-color: #333;
                selection-background-color: #d4af37;
                selection-color: black;
                outline: none;
            }
        """)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    storekeeper_window = StorekeeperApp(1)
    storekeeper_window.show()
    sys.exit(app.exec())